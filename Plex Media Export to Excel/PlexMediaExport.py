#!/usr/bin/env python3
################################################################################
# Plex Media Export Script - Optimized Version
# 
# Purpose: Generate a comprehensive Excel report of Plex Media Server content,
#           including movies and TV shows with TVMaze integration.
#
# Features:
# - Automatically loads environment variables from a .env file
# - Customizable movie & TV show export fields via .env file
# - Optimized field fetching: only processes requested fields from Plex objects
# - Parallel processing for movie data collection
# - Session management for improved API performance
# - Memory optimized Excel generation
# - Cached TVMaze lookups
# - Resolution-based highlighting (4K Green, 1080p White, <1080p Yellow)
# - TV show completion tracking
# - Sortable and filterable Excel headers (using TableStyleMedium9 for all sheets with row stripes)
################################################################################

# Standard library imports
from datetime import datetime  # For timestamping the output file and formatting dates
from typing import Dict, List, Optional, Union, Tuple # For type hinting, improving code readability and maintainability
import sys  # For system-specific parameters and functions, like exiting the script
from concurrent.futures import ThreadPoolExecutor  # For parallel processing of tasks (like fetching movie details)
from functools import lru_cache  # For caching results of functions (like TVMaze API calls)
import os  # For interacting with the operating system (like accessing environment variables and file paths)

# Third-party library imports
from dotenv import load_dotenv  # For loading environment variables from a .env file
from plexapi.server import PlexServer  # The main class for interacting with a Plex Media Server
import pandas as pd  # For data manipulation and analysis, used here for creating DataFrames for movies
from openpyxl import Workbook  # For creating and manipulating Excel (.xlsx) files
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment  # For styling cells in Excel
from openpyxl.utils import get_column_letter  # Utility to convert column index to letter (e.g., 1 -> A)
from openpyxl.worksheet.table import Table, TableStyleInfo  # For creating formatted tables within Excel worksheets
from requests import Session  # For making HTTP requests with connection pooling and session persistence
import requests  # For requests.exceptions, to catch specific HTTP request errors

# --- Load Environment Variables ---
# This function call loads variables defined in a .env file in the script's directory (or parent directories)
# into environment variables, making them accessible via os.environ.get().
# This is useful for storing sensitive information like API keys or URLs outside the script.
load_dotenv()

# --- Configuration ---
# Define keys for environment variables to ensure consistency and avoid typos when accessing them.
PLEX_URL_ENV_VAR = 'PLEX_URL'
PLEX_TOKEN_ENV_VAR = 'PLEX_TOKEN'
PLEX_EXPORT_DIR_ENV_VAR = 'PLEX_EXPORT_DIR'
PLEX_MOVIE_EXPORT_FIELDS_ENV_VAR = 'PLEX_MOVIE_EXPORT_FIELDS'
PLEX_SHOW_EXPORT_FIELDS_ENV_VAR = 'PLEX_SHOW_EXPORT_FIELDS'

# Retrieve Plex server URL and token from environment variables.
# These are essential for connecting to the Plex server.
PLEX_URL = os.environ.get(PLEX_URL_ENV_VAR)
PLEX_TOKEN = os.environ.get(PLEX_TOKEN_ENV_VAR)
# Retrieve the desired output directory for the Excel file. Defaults to the current directory (".") if not set.
PLEX_EXPORT_DIR = os.environ.get(PLEX_EXPORT_DIR_ENV_VAR, ".")
# Define the base URL for the TVMaze API.
TVMAZE_API = 'https://api.tvmaze.com'

# --- Movie Field Configuration ---
# Defines a comprehensive list of all movie fields that this script is capable of processing.
# This list is used to validate user-defined fields from the .env file.
ALL_POSSIBLE_MOVIE_FIELDS = [
    'Title', 'Year', 'Studio', 'ContentRating',
    'Video Resolution', 'Bitrate (kbps)', 'File Path', 'Container', 'Duration (min)',
    'AddedAt', 'LastViewedAt', 'OriginallyAvailableAt', 'Summary', 'Tagline',
    'AudienceRating', 'Rating', 'Collections', 'Genres', 'Labels',
    'AspectRatio', 'AudioChannels', 'AudioCodec', 'VideoCodec', 'VideoFrameRate',
    'Height', 'Width', 'ViewCount', 'SkipCount'
]
# Defines the default set of movie fields to be exported if the user doesn't specify any.
DEFAULT_MOVIE_FIELDS = [
    'Title', 'Year', 'Studio', 'ContentRating',
    'Video Resolution', 'Bitrate (kbps)', 'File Path', 'Container', 'Duration (min)'
]
# Retrieve the user's preferred movie export fields from the .env file.
user_defined_movie_fields_str = os.environ.get(PLEX_MOVIE_EXPORT_FIELDS_ENV_VAR)
if user_defined_movie_fields_str:
    # If the user provided a string, split it by commas, strip whitespace from each field.
    temp_selected_movie_fields = [field.strip() for field in user_defined_movie_fields_str.split(',') if field.strip()]
    # Filter the user's list against ALL_POSSIBLE_MOVIE_FIELDS to ensure only valid fields are used
    # and to maintain a consistent order if the user's order differs (though this implementation uses user's order if valid).
    SELECTED_MOVIE_FIELDS = [field for field in temp_selected_movie_fields if field in ALL_POSSIBLE_MOVIE_FIELDS]
    if not SELECTED_MOVIE_FIELDS:
        # If the user's string was all invalid fields or just commas, fall back to defaults.
        print(f"Warning: '{PLEX_MOVIE_EXPORT_FIELDS_ENV_VAR}' in .env contains no valid fields or is empty. Using default movie fields.")
        SELECTED_MOVIE_FIELDS = DEFAULT_MOVIE_FIELDS[:] # Use a copy of the default list
else:
    # If the environment variable is not set, use the default fields.
    SELECTED_MOVIE_FIELDS = DEFAULT_MOVIE_FIELDS[:] # Use a copy
# Print the final list of movie fields that will be exported.
print(f"Selected movie fields for export: {', '.join(SELECTED_MOVIE_FIELDS)}")

# --- TV Show Field Configuration ---
# Defines all possible base metadata fields for TV shows that this script can process.
# Season/episode specific details are handled separately.
ALL_POSSIBLE_SHOW_FIELDS = [
    'Title', 'Year', 'Studio', 'ContentRating', 'Summary', 'Tagline',
    'AddedAt', 'LastViewedAt', 'OriginallyAvailableAt',
    'AudienceRating', 'Rating', 'Collections', 'Genres', 'Labels',
    'ViewCount', 'SkipCount'
]
# Defines the default set of TV show base fields to export.
DEFAULT_SHOW_FIELDS = [
    'Title', 'Year', 'Studio', 'ContentRating', 'Summary'
]
# Retrieve user's preferred TV show export fields from .env.
user_defined_show_fields_str = os.environ.get(PLEX_SHOW_EXPORT_FIELDS_ENV_VAR)
if user_defined_show_fields_str:
    # Process the user's string similar to how movie fields are processed.
    temp_selected_show_fields = [field.strip() for field in user_defined_show_fields_str.split(',') if field.strip()]
    SELECTED_SHOW_FIELDS = [field for field in temp_selected_show_fields if field in ALL_POSSIBLE_SHOW_FIELDS]
    if not SELECTED_SHOW_FIELDS:
        print(f"Warning: '{PLEX_SHOW_EXPORT_FIELDS_ENV_VAR}' in .env contains no valid fields or is empty. Using default show fields.")
        SELECTED_SHOW_FIELDS = DEFAULT_SHOW_FIELDS[:]
else:
    SELECTED_SHOW_FIELDS = DEFAULT_SHOW_FIELDS[:]
# Print the final list of TV show base fields for export.
print(f"Selected TV show fields for export: {', '.join(SELECTED_SHOW_FIELDS)}")

# --- Initialization ---
# Create a global requests.Session object. This allows for connection pooling,
# which can improve performance when making multiple HTTP requests to the same host (TVMaze API).
session = Session()

# Define a dictionary of styles to be used for formatting Excel cells.
# This centralizes style definitions for consistency and easier modification.
STYLES = {
    'fills': {  # Define various fill patterns (background colors)
        'gray': PatternFill(patternType='solid', fgColor='D3D3D3'),  # For non-existent TV seasons
        'green': PatternFill(patternType='solid', fgColor='90EE90'), # For complete TV series/seasons
        'red': PatternFill(patternType='solid', fgColor='FFB6B6'),   # For incomplete TV series/seasons
        'yellow': PatternFill(patternType='solid', fgColor='FFFFCC'),# For low-resolution movies (<1080p)
        '4k_specific_light_green': PatternFill(patternType='solid', fgColor='77b190'), # For 4K movies (dark green)
        'white': PatternFill(patternType='solid', fgColor='FFFFFF')  # Explicit white fill (for 1080p movies)
    },
    'borders': { # Define border styles
        'thin': Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')),
        'header': Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thick')) # Thicker bottom border for headers
    },
    'fonts': { # Define font styles
        'bold': Font(bold=True) # Bold font for headers
    },
    'alignments': { # Define text alignment styles
        'center': Alignment(horizontal='center', vertical='center', wrap_text=False), # Center alignment, no text wrapping
        'left': Alignment(horizontal='left', vertical='center', wrap_text=False),   # Left alignment, no text wrapping
        'summary': Alignment(horizontal='left', vertical='top', wrap_text=True)     # Left-top alignment with text wrapping (for summaries)
    }
}

def connect_to_plex() -> Optional[PlexServer]:
    """
    Establishes a connection to the Plex Media Server using credentials from environment variables.

    Returns:
        Optional[PlexServer]: A PlexServer instance if connection is successful, None otherwise.
    """
    # Check if PLEX_URL and PLEX_TOKEN are set.
    if not PLEX_URL or not PLEX_TOKEN:
        print(f"Error: Plex URL ({PLEX_URL_ENV_VAR}) and Token ({PLEX_TOKEN_ENV_VAR}) must be set.")
        return None
    try:
        # Create a new session for Plex API calls. This is separate from the global session for TVMaze.
        plex_session = Session()
        # Note: SSL verification can be disabled here if needed (e.g., for self-signed certs on a local network),
        # but it's generally not recommended for security reasons.
        # Example:
        # from urllib3.exceptions import InsecureRequestWarning
        # requests.packages.urllib3.disable_warnings(category=InsecureRequestWarning)
        # plex_session.verify = False

        print(f"Connecting to Plex server at {PLEX_URL}...")
        # Initialize the PlexServer object with the URL, token, and session.
        server = PlexServer(PLEX_URL, PLEX_TOKEN, session=plex_session)
        # A simple test to confirm connection by fetching the server's friendly name.
        print(f"Successfully connected to Plex server: {server.friendlyName}")
        return server
    except Exception as e:
        # Catch any exception during connection and print an error message.
        print(f"Failed to connect to Plex server: {e}")
        return None

@lru_cache(maxsize=256) # Decorator to cache results of this function.
def get_tvmaze_show_info(search_term: str, is_imdb_id: bool = False) -> Optional[Dict]:
    """
    Fetches TV show information (seasons, episode counts) from the TVMaze API.
    Results are cached to avoid redundant API calls for the same show.

    Args:
        search_term (str): The name of the TV show or its IMDB ID.
        is_imdb_id (bool): True if search_term is an IMDB ID, False if it's a show title.

    Returns:
        Optional[Dict]: A dictionary containing total seasons and episode counts per season,
                        or None if the show is not found or an error occurs.
    """
    try:
        show_id = None
        params = {}
        api_url = "" # Initialize api_url

        # Determine the TVMaze API endpoint and parameters based on whether searching by IMDB ID or title.
        if is_imdb_id:
            api_url, params = f"{TVMAZE_API}/lookup/shows", {'imdb': search_term}
        else:
            api_url, params = f"{TVMAZE_API}/search/shows", {'q': search_term}

        # Make the GET request to TVMaze API using the global session.
        response = session.get(api_url, params=params, timeout=10) # Added a 10-second timeout
        response.raise_for_status() # Raise an HTTPError for bad status codes (4xx or 5xx)
        
        json_response = response.json() # Parse the JSON response

        # Extract the show ID from the response.
        if is_imdb_id: # /lookup/shows returns a single show object or null
            if json_response: # Check if response is not null
                show_id = json_response.get('id')
        elif json_response and isinstance(json_response, list) and len(json_response) > 0: # /search/shows returns a list
            # Assume the first result is the most relevant one.
            show_id = json_response[0].get('show', {}).get('id')

        # If a show ID was found, fetch episode details.
        if show_id:
            episodes_response = session.get(f"{TVMAZE_API}/shows/{show_id}/episodes", timeout=10)
            episodes_response.raise_for_status()
            episodes = episodes_response.json()
            
            seasons_data = {}
            # If no episodes are listed, return 0 seasons.
            if not episodes:
                return {'total_seasons': 0, 'seasons': {}}
            
            # Aggregate episode counts by season number.
            for episode in episodes:
                season_num = episode.get('season')
                if season_num is None: # Skip episodes with no season number (e.g., some specials)
                    continue
                seasons_data.setdefault(season_num, {'total_episodes': 0})['total_episodes'] += 1
            
            # If no valid season data was processed (e.g., all episodes lacked season numbers).
            if not seasons_data:
                return {'total_seasons': 0, 'seasons': {}}

            # Return the total number of seasons and the per-season episode counts.
            return {'total_seasons': max(seasons_data.keys(), default=0), 'seasons': seasons_data}
            
    # Handle specific exceptions that might occur during API requests or JSON parsing.
    except requests.exceptions.RequestException as e:
        print(f"TVMaze API request error for '{search_term}': {e}")
    except ValueError as e: # Includes JSONDecodeError
        print(f"TVMaze API JSON decode error for '{search_term}': {e}")
    except Exception as e: # Catch any other unexpected errors
        print(f"Unexpected error fetching TVMaze data for '{search_term}': {e}")
    return None # Return None if any error occurs

def format_plex_datetime(dt_obj) -> str:
    """
    Formats a datetime object (often from Plex attributes) into an ISO 8601 string.
    Handles cases where the datetime object might be None.

    Args:
        dt_obj: The datetime object to format.

    Returns:
        str: The formatted datetime string or 'N/A' if input is None.
    """
    if dt_obj is None:
        return 'N/A'
    try:
        # Check if it's a datetime or pandas Timestamp object before calling isoformat.
        return dt_obj.isoformat() if isinstance(dt_obj, (datetime, pd.Timestamp)) else str(dt_obj)
    except AttributeError: # If it's not a datetime object but something else that doesn't have isoformat
        return str(dt_obj)

def format_plex_list(plex_list_attr) -> str:
    """
    Formats a list of Plex Tag objects (e.g., genres, collections, labels)
    into a comma-separated string of their 'tag' or 'title' attributes.

    Args:
        plex_list_attr: A list of Plex Tag-like objects.

    Returns:
        str: A comma-separated string of tags/titles, or 'N/A' if the list is empty/None.
    """
    if plex_list_attr is None or not plex_list_attr: # Check if None or empty list
        return 'N/A'
    try:
        # Plex Tag objects usually have a 'tag' attribute (for Genre, Director, etc.)
        # or a 'title' attribute (for Collection). This attempts to get the most relevant one.
        return ', '.join(
            item.tag if hasattr(item, 'tag') else
            item.title if hasattr(item, 'title') else
            str(item) # Fallback to string representation if neither 'tag' nor 'title' exists
            for item in plex_list_attr
        )
    except Exception: # Catch any error during list processing
        return 'Error processing list'

def process_movie(movie) -> Dict:
    """
    Processes a single Plex movie object and extracts data for the fields
    specified in SELECTED_MOVIE_FIELDS.

    Args:
        movie: A Plex movie object.

    Returns:
        Dict: A dictionary containing the extracted movie data.
    """
    movie_data = {} # Initialize an empty dictionary to store data for selected fields
    media = None    # Will hold the primary media object for the movie
    parts = None    # Will hold the primary part object from the media

    # Determine if any media-dependent fields are selected for export.
    # If so, fetch the media and parts objects once to avoid redundant lookups.
    media_dependent_fields = [
        'Video Resolution', 'Bitrate (kbps)', 'File Path', 'Container',
        'AspectRatio', 'AudioChannels', 'AudioCodec', 'VideoCodec',
        'VideoFrameRate', 'Height', 'Width'
    ]
    if any(field in SELECTED_MOVIE_FIELDS for field in media_dependent_fields):
        media = movie.media[0] if movie.media else None # Get the first media item
        if media:
            parts = media.parts[0] if media.parts else None # Get the first part of that media item

    # Iterate through only the fields selected by the user for export.
    for field_name in SELECTED_MOVIE_FIELDS:
        value = 'N/A' # Default value for any field
        try:
            # Retrieve data for each selected field based on its name.
            # Uses hasattr to check for attribute existence before accessing to prevent errors.
            if field_name == 'Title':
                if hasattr(movie, 'title'): value = movie.title
            elif field_name == 'Year':
                if hasattr(movie, 'year'): value = movie.year
            elif field_name == 'Studio':
                if hasattr(movie, 'studio') and movie.studio is not None: value = movie.studio
            elif field_name == 'ContentRating':
                if hasattr(movie, 'contentRating') and movie.contentRating is not None: value = movie.contentRating
            elif field_name == 'Duration (min)':
                if hasattr(movie, 'duration') and movie.duration is not None: value = round(movie.duration / 60000) # Convert ms to minutes
            elif field_name == 'AddedAt':
                if hasattr(movie, 'addedAt'): value = format_plex_datetime(movie.addedAt)
            elif field_name == 'LastViewedAt':
                if hasattr(movie, 'lastViewedAt'): value = format_plex_datetime(movie.lastViewedAt)
            elif field_name == 'OriginallyAvailableAt':
                if hasattr(movie, 'originallyAvailableAt'): value = format_plex_datetime(movie.originallyAvailableAt)
            elif field_name == 'Summary':
                if hasattr(movie, 'summary') and movie.summary is not None: value = movie.summary
            elif field_name == 'Tagline':
                if hasattr(movie, 'tagline') and movie.tagline is not None: value = movie.tagline
            elif field_name == 'AudienceRating':
                if hasattr(movie, 'audienceRating') and movie.audienceRating is not None: value = movie.audienceRating
            elif field_name == 'Rating': # Typically critic rating
                if hasattr(movie, 'rating') and movie.rating is not None: value = movie.rating
            elif field_name == 'Collections':
                if hasattr(movie, 'collections'): value = format_plex_list(movie.collections)
            elif field_name == 'Genres':
                if hasattr(movie, 'genres'): value = format_plex_list(movie.genres)
            elif field_name == 'Labels':
                if hasattr(movie, 'labels'): value = format_plex_list(movie.labels)
            elif field_name == 'ViewCount':
                if hasattr(movie, 'viewCount') and movie.viewCount is not None: value = movie.viewCount
                else: value = 0 # Default to 0 if not present
            elif field_name == 'SkipCount':
                if hasattr(movie, 'skipCount') and movie.skipCount is not None: value = movie.skipCount
                else: value = 0 # Default to 0
            # Media-dependent fields are processed only if 'media' object was successfully retrieved.
            elif media: 
                if field_name == 'Video Resolution':
                    res_val_raw = getattr(media, 'videoResolution', None)
                    # Ensure robust handling: if raw value is None, empty, or just whitespace, set to 'Unknown'.
                    value = str(res_val_raw).strip() if res_val_raw and isinstance(res_val_raw, str) and res_val_raw.strip() else 'Unknown'
                elif field_name == 'Bitrate (kbps)':
                    br_val = getattr(media, 'bitrate', None)
                    value = br_val if br_val is not None and str(br_val).strip() else 'Unknown'
                elif field_name == 'Container':
                    con_val = getattr(media, 'container', None)
                    value = con_val if con_val and str(con_val).strip() else 'Unknown'
                elif field_name == 'AspectRatio':
                    if hasattr(media, 'aspectRatio') and media.aspectRatio is not None: value = media.aspectRatio
                elif field_name == 'AudioChannels':
                    if hasattr(media, 'audioChannels') and media.audioChannels is not None: value = media.audioChannels
                elif field_name == 'AudioCodec':
                    if hasattr(media, 'audioCodec') and media.audioCodec is not None: value = media.audioCodec
                elif field_name == 'VideoCodec':
                    if hasattr(media, 'videoCodec') and media.videoCodec is not None: value = media.videoCodec
                elif field_name == 'VideoFrameRate':
                    if hasattr(media, 'videoFrameRate') and media.videoFrameRate is not None: value = media.videoFrameRate
                elif field_name == 'Height':
                    if hasattr(media, 'height') and media.height is not None: value = media.height
                elif field_name == 'Width':
                    if hasattr(media, 'width') and media.width is not None: value = media.width
                elif field_name == 'File Path':
                    if parts and hasattr(parts, 'file'): value = parts.file
                    else: value = 'Unknown'
        except Exception as e:
            # If any error occurs while processing a field, log it and set value to 'Error Processing Field'.
            print(f"Error processing field '{field_name}' for movie '{getattr(movie, 'title', 'Unknown Title')}': {e}")
            value = 'Error Processing Field'
        movie_data[field_name] = value # Store the processed value (or 'N/A' or error string)

    return movie_data

def get_movie_details(movies: List) -> List[Dict]:
    """
    Retrieves details for a list of Plex movie objects in parallel using a ThreadPoolExecutor.

    Args:
        movies (List): A list of Plex movie objects.

    Returns:
        List[Dict]: A list of dictionaries, where each dictionary contains details for a movie.
    """
    # Determine a reasonable number of worker threads.
    num_workers = min(10, (os.cpu_count() or 1) + 4) # Max 10 workers, or CPU count + 4
    # Use ThreadPoolExecutor to process movies in parallel.
    # executor.map applies process_movie to each movie in the list.
    with ThreadPoolExecutor(max_workers=num_workers) as executor:
        # filter(None, ...) removes any None results if process_movie were to return None (it currently doesn't).
        return list(filter(None, executor.map(process_movie, movies)))

def process_show_metadata(show_obj) -> Dict:
    """
    Processes a single Plex TV show object and extracts base metadata for the fields
    specified in SELECTED_SHOW_FIELDS.

    Args:
        show_obj: A Plex TV show object.

    Returns:
        Dict: A dictionary containing the extracted TV show metadata.
    """
    show_metadata = {} # Initialize an empty dictionary for selected fields
    # Iterate through only the fields selected by the user for TV show export.
    for field_name in SELECTED_SHOW_FIELDS:
        value = 'N/A' # Default value
        try:
            # Retrieve data for each selected field.
            if field_name == 'Title':
                if hasattr(show_obj, 'title'): value = show_obj.title
            elif field_name == 'Year':
                if hasattr(show_obj, 'year'): value = show_obj.year
            elif field_name == 'Studio':
                if hasattr(show_obj, 'studio') and show_obj.studio is not None: value = show_obj.studio
            elif field_name == 'ContentRating':
                if hasattr(show_obj, 'contentRating') and show_obj.contentRating is not None: value = show_obj.contentRating
            elif field_name == 'Summary':
                if hasattr(show_obj, 'summary') and show_obj.summary is not None: value = show_obj.summary
            elif field_name == 'Tagline':
                if hasattr(show_obj, 'tagline') and show_obj.tagline is not None: value = show_obj.tagline
            elif field_name == 'AddedAt':
                if hasattr(show_obj, 'addedAt'): value = format_plex_datetime(show_obj.addedAt)
            elif field_name == 'LastViewedAt':
                if hasattr(show_obj, 'lastViewedAt'): value = format_plex_datetime(show_obj.lastViewedAt)
            elif field_name == 'OriginallyAvailableAt':
                if hasattr(show_obj, 'originallyAvailableAt'): value = format_plex_datetime(show_obj.originallyAvailableAt)
            elif field_name == 'AudienceRating':
                if hasattr(show_obj, 'audienceRating') and show_obj.audienceRating is not None: value = show_obj.audienceRating
            elif field_name == 'Rating':
                if hasattr(show_obj, 'rating') and show_obj.rating is not None: value = show_obj.rating
            elif field_name == 'Collections':
                if hasattr(show_obj, 'collections'): value = format_plex_list(show_obj.collections)
            elif field_name == 'Genres':
                if hasattr(show_obj, 'genres'): value = format_plex_list(show_obj.genres)
            elif field_name == 'Labels':
                if hasattr(show_obj, 'labels'): value = format_plex_list(show_obj.labels)
            elif field_name == 'ViewCount':
                if hasattr(show_obj, 'viewCount') and show_obj.viewCount is not None: value = show_obj.viewCount
                else: value = 0
            elif field_name == 'SkipCount':
                if hasattr(show_obj, 'skipCount') and show_obj.skipCount is not None: value = show_obj.skipCount
                else: value = 0
        except Exception as e:
            print(f"Error processing field '{field_name}' for show '{getattr(show_obj, 'title', 'Unknown Title')}': {e}")
            value = 'Error Processing Field'
        show_metadata[field_name] = value
    return show_metadata

def get_show_details(shows: List) -> Tuple[List[Dict], int]:
    """
    Retrieves detailed information for a list of TV shows, including base metadata,
    Plex season/episode counts, and TVMaze information.

    Args:
        shows (List): A list of Plex TV show objects.

    Returns:
        Tuple[List[Dict], int]: A tuple containing:
            - A list of dictionaries, each with details for a TV show.
            - An integer representing the maximum number of seasons found across all shows.
    """
    processed_shows_data = []
    max_seasons_overall = 0 # Track the highest season number encountered for column generation.
    
    # Iterate through each show object from Plex.
    for i, show_obj in enumerate(shows): 
        print(f"Processing TV Show: {show_obj.title} ({i+1}/{len(shows)})")
        
        # Get the base metadata for the show using the selected fields.
        show_metadata = process_show_metadata(show_obj) 
        
        # Attempt to find IMDB ID for TVMaze lookup.
        imdb_id_full = next((guid.id for guid in show_obj.guids if guid.id.startswith('imdb://')), None)
        imdb_id = imdb_id_full.split('imdb://')[-1] if imdb_id_full else None
        
        tvmaze_info = None
        if imdb_id: # Try TVMaze lookup with IMDB ID first.
            tvmaze_info = get_tvmaze_show_info(imdb_id, is_imdb_id=True)
        if not tvmaze_info: # If IMDB lookup fails or no ID, try with show title.
            search_title = show_obj.originalTitle if show_obj.originalTitle else show_obj.title
            tvmaze_info = get_tvmaze_show_info(search_title)
        
        # Update the overall maximum number of seasons if TVMaze info is found.
        if tvmaze_info and tvmaze_info.get('total_seasons') is not None:
            max_seasons_overall = max(max_seasons_overall, tvmaze_info['total_seasons'])
        elif not tvmaze_info:
            print(f"  Could not find TVMaze info for: {show_obj.title} (IMDB: {imdb_id})")
            
        # Get season and episode counts from Plex.
        plex_seasons_data = {}
        try:
            for s in show_obj.seasons():
                if s.seasonNumber is not None: # Only process seasons with a valid number
                    try:
                        season_num_int = int(s.seasonNumber)
                        plex_seasons_data[season_num_int] = {
                            'episodes_in_plex': len(s.episodes()),
                            'season_number': season_num_int
                        }
                    except ValueError: # Handle cases where seasonNumber might not be a valid integer
                        print(f"  Skipping invalid season number '{s.seasonNumber}' for show '{show_obj.title}'")
        except Exception as e:
            print(f"  Error fetching Plex season details for {show_obj.title}: {e}")
            
        # Combine the base metadata, Plex season data, and TVMaze info into one dictionary for the show.
        combined_show_data = {**show_metadata, 'seasons_in_plex': plex_seasons_data, 'tvmaze_info': tvmaze_info}
        processed_shows_data.append(combined_show_data)
        
    return processed_shows_data, max_seasons_overall

def apply_cell_styling(cell, is_header: bool = False, alignment_key: str = 'center', fill_pattern=None):
    """
    Applies predefined border, alignment, font (if header), and fill styles to an Excel cell.

    Args:
        cell: The openpyxl cell object to style.
        is_header (bool): True if the cell is a header cell, False otherwise.
        alignment_key (str): The key for the desired alignment style from STYLES['alignments'].
        fill_pattern: The PatternFill object to apply as background, or None for no fill.
    """
    # Apply border style (different for header vs. data cells).
    cell.border = STYLES['borders']['header' if is_header else 'thin']
    # Apply alignment style, defaulting to 'left' if the key is not found.
    alignment_style = STYLES['alignments'].get(alignment_key, STYLES['alignments']['left'])
    cell.alignment = alignment_style
    # Apply bold font if it's a header cell.
    if is_header:
        cell.font = STYLES['fonts']['bold']
    # Apply fill pattern if one is provided.
    if fill_pattern:
        cell.fill = fill_pattern

def create_table(ws, table_name: str, data_range: str, style_name: str = 'TableStyleMedium9'):
    """
    Creates an Excel table on the given worksheet with specified name, range, and style.
    Ensures table names are sanitized (no spaces) and enables row stripes by default.

    Args:
        ws: The openpyxl worksheet object.
        table_name (str): The desired name for the table (will be sanitized).
        data_range (str): The cell range for the table (e.g., "A1:G100").
        style_name (str): The name of the built-in Excel table style to apply.

    Returns:
        Table: The created openpyxl Table object, or None if an error occurs.
    """
    # Sanitize table name: Excel table names cannot contain spaces or certain characters,
    # and must not start with a number.
    sanitized_table_name = table_name.replace(" ", "_").replace("-", "_")
    if not sanitized_table_name or not sanitized_table_name[0].isalpha():
        sanitized_table_name = "T_" + sanitized_table_name # Prepend "T_" if name is invalid
        
    try:
        # Create the Table object.
        table = Table(displayName=sanitized_table_name, ref=data_range)
        # Define the table style information.
        # showRowStripes=True enables alternating row colors from the chosen style.
        # Set to False if you want custom cell fills to always take precedence without interference.
        table.tableStyleInfo = TableStyleInfo(
            name=style_name, showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False
        )
        # Add the table to the worksheet.
        ws.add_table(table)
        return table
    except Exception as e:
        print(f"Error creating table '{sanitized_table_name}': {e}")
        return None

def auto_adjust_columns(ws, min_width: int = 10, max_width: int = 50, wrap_text_columns: List[str] = None):
    """
    Automatically adjusts column widths on a worksheet based on content length.
    Allows specifying columns that should rely on text wrapping instead of auto-sizing.

    Args:
        ws: The openpyxl worksheet object.
        min_width (int): The minimum width for any column.
        max_width (int): The maximum width for auto-sized columns.
        wrap_text_columns (List[str], optional): A list of header names for columns
                                                 that should be set to max_width and rely on wrap text.
    """
    if wrap_text_columns is None: # Initialize to empty list if not provided
        wrap_text_columns = []
        
    header_row = [cell.value for cell in ws[1]] # Get header values from the first row of the worksheet

    # Iterate through each column in the worksheet.
    for col_idx, column_cells in enumerate(ws.columns):
        column_letter = column_cells[0].column_letter # Get the column letter (e.g., 'A', 'B')
        current_header = header_row[col_idx] if col_idx < len(header_row) else ""

        # If the current column is designated for text wrapping, set its width to max_width and skip auto-sizing.
        if current_header in wrap_text_columns:
            ws.column_dimensions[column_letter].width = max_width 
            continue
            
        try:
            # Calculate the maximum length of content in the current column.
            base_length = max(len(str(cell.value)) for cell in column_cells if cell.value is not None)
            length = base_length + 2 # Add a little padding
            # Constrain the width between min_width and max_width.
            final_width = max(min_width, min(length, max_width))
            ws.column_dimensions[column_letter].width = final_width
        except ValueError: # Handles case where column might be entirely empty, max() on empty sequence
             ws.column_dimensions[column_letter].width = min_width
        except Exception as e: # Catch any other errors during column adjustment
            print(f"Error auto-adjusting column {column_letter}: {e}. Setting default width.")
            ws.column_dimensions[column_letter].width = min_width

def create_movies_worksheet(section_name: str, wb: Workbook, movie_list: List[Dict]):
    """
    Creates and populates a worksheet for movies.

    Args:
        section_name (str): The name of the Plex library section (used for sheet name).
        wb (Workbook): The openpyxl Workbook object.
        movie_list (List[Dict]): A list of dictionaries, each containing movie data.
    """
    if not movie_list: # If no movies were found/processed for this section
        print(f"No movies found in section '{section_name}'. Skipping sheet creation.")
        return

    # Sanitize the section name to be a valid Excel sheet name (max 31 chars, no invalid chars).
    safe_sheet_name = "".join(c if c.isalnum() or c in " _-" else "_" for c in section_name[:31])
    ws = wb.create_sheet(safe_sheet_name) # Create the new worksheet
    ws.freeze_panes = 'A2' # Freeze the header row (row 1) and first column (A) is not frozen here.

    # The movie_list already contains dictionaries with only SELECTED_MOVIE_FIELDS due to process_movie.
    df_data = movie_list 
    if not df_data: # Should not happen if movie_list is not empty, but as a safeguard.
        print(f"No data to process for movies in section '{section_name}' after field selection.")
        return

    # Create a Pandas DataFrame from the movie data, ensuring columns are in the order of SELECTED_MOVIE_FIELDS.
    df = pd.DataFrame(df_data, columns=SELECTED_MOVIE_FIELDS) 
    if df.empty: # If DataFrame is empty after creation (e.g., if df_data was empty)
        print(f"DataFrame is empty for section '{section_name}' after field selection. Skipping sheet content.")
        return
    
    # Determine the field to sort by: 'Title' if selected, otherwise the first selected field.
    sort_by_field = 'Title' if 'Title' in SELECTED_MOVIE_FIELDS else (SELECTED_MOVIE_FIELDS[0] if SELECTED_MOVIE_FIELDS else None)
    if sort_by_field:
        df = df.sort_values(by=sort_by_field, ascending=True) # Sort the DataFrame
    
    headers = SELECTED_MOVIE_FIELDS # The headers are the selected fields

    # Write and style the header row.
    for col_idx, header_text in enumerate(headers, 1):
        apply_cell_styling(ws.cell(row=1, column=col_idx, value=header_text), is_header=True, alignment_key='center')
    
    # --- Debugging Movie Fills ---
    # This section was previously used for debugging and can be re-enabled if needed.
    # print(f"\n--- Debugging Movie Fills for Section: {section_name} ---") 
    
    # Iterate through each row of movie data in the DataFrame.
    for row_idx, row_data_tuple in enumerate(df.itertuples(index=False), 2): # Start from Excel row 2
        row_fill = None # Default to no specific fill for the row.
        
        # Determine row fill color based on 'Video Resolution' if that field is selected for export.
        if 'Video Resolution' in SELECTED_MOVIE_FIELDS:
            try:
                # Access the 'Video Resolution' value from the named tuple.
                # Pandas DataFrame's itertuples() replaces spaces in column names with underscores for attribute access.
                # However, since we construct the DataFrame with SELECTED_MOVIE_FIELDS, the tuple elements are in that order.
                # Accessing by index is more robust if column names are complex.
                res_col_idx = SELECTED_MOVIE_FIELDS.index('Video Resolution')
                raw_resolution_from_tuple = row_data_tuple[res_col_idx]
                
                resolution_val = str(raw_resolution_from_tuple).strip().lower() # Process for comparison

                # Apply fill based on resolution.
                if resolution_val in ['4k', 'uhd', '2160p']: row_fill = STYLES['fills']['4k_specific_light_green']
                elif resolution_val in ['1080', '1080p']: row_fill = STYLES['fills']['white'] 
                elif resolution_val in ['sd', '480', '480p', '576', '576p', '720', '720p', 'dvd', 'pal', 'ntsc']:
                    row_fill = STYLES['fills']['yellow']
            except (ValueError, IndexError, AttributeError) as e: 
                # Error accessing or processing resolution, so no specific fill.
                # print(f"  DEBUG: Error determining fill for row {row_idx-1}: {e}") # Optional debug
                pass 
        
        # --- Debug Print (can be re-enabled) ---
        # movie_title_debug = getattr(row_data_tuple, 'Title', 'N/A') if 'Title' in SELECTED_MOVIE_FIELDS else 'N/A'
        # print(f"Movie: {movie_title_debug}, ProcessedRes='{resolution_val if 'Video Resolution' in SELECTED_MOVIE_FIELDS else 'N/A'}', DeterminedFill='{row_fill}'")
        # if row_fill:
        #     print(f"  Fill Type: {type(row_fill)}, FG Color: {row_fill.fgColor.rgb if hasattr(row_fill, 'fgColor') and row_fill.fgColor else 'N/A'}")
        
        # Write data for each cell in the current row and apply styles.
        for col_idx, value in enumerate(row_data_tuple, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            # Ensure value is appropriate for Excel cell (numeric or string).
            cell.value = value if isinstance(value, (int, float)) and not pd.isna(value) else str(value if not pd.isna(value) else '')
            
            current_header = headers[col_idx-1] # Get current header name
            # Determine alignment based on header.
            alignment_key = 'summary' if current_header in ['Summary', 'Tagline'] else \
                            'left' if current_header in ['Title', 'File Path', 'Studio', 'Collections', 'Genres', 'Labels'] else 'center'
            apply_cell_styling(cell, alignment_key=alignment_key, fill_pattern=row_fill) # Apply determined row fill
            
    # Create an Excel table for the movie data.
    last_row, last_col_letter = len(df) + 1, get_column_letter(len(headers))
    if last_row > 1 and len(headers) > 0:
        create_table(ws, f'{"".join(c if c.isalnum() else "_" for c in safe_sheet_name)}_Table', f"A1:{last_col_letter}{last_row}", style_name='TableStyleMedium9')
    # Adjust column widths.
    auto_adjust_columns(ws, max_width=60, wrap_text_columns=['Summary', 'Tagline', 'File Path', 'Collections', 'Genres', 'Labels'])

def create_tv_shows_worksheet(section_name: str, wb: Workbook, shows_data_full: List[Dict], max_seasons_overall: int):
    """
    Creates and populates a worksheet for TV shows.

    Args:
        section_name (str): The name of the Plex library section.
        wb (Workbook): The openpyxl Workbook object.
        shows_data_full (List[Dict]): A list of dictionaries, each containing full data for a TV show.
        max_seasons_overall (int): The maximum number of seasons found across all shows for header generation.
    """
    if not shows_data_full: print(f"No TV shows found in section '{section_name}'. Skipping sheet creation."); return
    safe_sheet_name = "".join(c if c.isalnum() or c in " _-" else "_" for c in section_name[:31])
    ws = wb.create_sheet(safe_sheet_name); 
    
    # Sort the full show data list before processing rows.
    # This ensures that the order of rows in Excel matches the sorted order.
    sort_by_field_show = 'Title' if 'Title' in SELECTED_SHOW_FIELDS else (SELECTED_SHOW_FIELDS[0] if SELECTED_SHOW_FIELDS else None)
    if sort_by_field_show:
        shows_data_full.sort(key=lambda x: str(x.get(sort_by_field_show, '')).lower()) # Case-insensitive sort

    # Define headers: selected base fields + completion status column + dynamic season columns.
    base_headers = SELECTED_SHOW_FIELDS[:] 
    completion_header = "Series Complete (Plex/TVMaze)"
    has_specials = any(0 in show_dict.get('seasons_in_plex', {}) for show_dict in shows_data_full)
    season_headers_list = []
    if has_specials: season_headers_list.append("S00") # Add "S00" for specials if any show has them.
    season_headers_list.extend([f"S{i:02d}" for i in range(1, max_seasons_overall + 1)]) # S01, S02, ...
    final_headers = base_headers + [completion_header] + season_headers_list
    
    # Freeze panes after the base headers and the completion status column.
    ws.freeze_panes = get_column_letter(len(base_headers) + 1) + '2' if len(base_headers) > 0 else 'A2'

    # Write and style the header row.
    for col_idx, header_text in enumerate(final_headers, 1):
        apply_cell_styling(ws.cell(row=1, column=col_idx, value=header_text), is_header=True, alignment_key='center')

    # Iterate through each show's data.
    for row_idx, show_info_full in enumerate(shows_data_full, 2): 
        current_col = 1 # Start from the first column for this row.
        
        # Write the selected base metadata fields for the show.
        for field_name in SELECTED_SHOW_FIELDS: 
            value = show_info_full.get(field_name, 'N/A') 
            cell = ws.cell(row=row_idx, column=current_col)
            cell.value = value if isinstance(value, (int, float)) and not pd.isna(value) else str(value if not pd.isna(value) else '')
            alignment_key = 'summary' if field_name in ['Summary', 'Tagline'] else \
                            'left' if field_name in ['Title', 'Studio', 'Collections', 'Genres', 'Labels'] else 'center'
            # No specific row_fill is applied here for base info; relies on table style or default cell style.
            apply_cell_styling(cell, alignment_key=alignment_key) 
            current_col += 1
            
        # Process and write the "Series Complete" status.
        tvmaze_info = show_info_full.get('tvmaze_info')
        plex_seasons_data = show_info_full.get('seasons_in_plex', {})
        series_status_text, series_fill = "N/A", None # Default values
        if tvmaze_info and tvmaze_info.get('total_seasons') is not None:
            # Calculate completion based on regular seasons (S1+).
            tvmaze_total_regular_seasons = sum(1 for s_num in tvmaze_info.get('seasons', {}) if s_num > 0)
            complete_plex_regular_seasons_count = 0
            for s_num, tvmaze_s_data in tvmaze_info.get('seasons', {}).items():
                if s_num == 0: continue # Skip specials (S00) for this metric
                plex_s_data = plex_seasons_data.get(s_num)
                if plex_s_data and tvmaze_s_data and \
                   plex_s_data.get('episodes_in_plex', 0) >= tvmaze_s_data.get('total_episodes', 0) and \
                   tvmaze_s_data.get('total_episodes', 0) > 0: # Ensure TVMaze season has episodes
                    complete_plex_regular_seasons_count += 1
            series_status_text = f"{complete_plex_regular_seasons_count}/{tvmaze_total_regular_seasons}"
            # Determine fill color for completion status.
            if tvmaze_total_regular_seasons == 0: series_fill = STYLES['fills']['gray']
            elif complete_plex_regular_seasons_count >= tvmaze_total_regular_seasons: series_fill = STYLES['fills']['green']
            elif complete_plex_regular_seasons_count > 0: series_fill = STYLES['fills']['red']
        else: # If no TVMaze info, show Plex season count vs '?'
            plex_regular_season_count = sum(1 for s_num in plex_seasons_data if s_num > 0)
            series_status_text = f"{plex_regular_season_count}/?"
            series_fill = STYLES['fills']['yellow'] if plex_regular_season_count > 0 else None
        
        status_cell = ws.cell(row=row_idx, column=current_col, value=series_status_text)
        apply_cell_styling(status_cell, fill_pattern=series_fill, alignment_key='center')
        current_col += 1
        
        # Process and write individual season episode counts (S00, S01, S02...).
        for season_header_text in season_headers_list:
            season_num_from_header = int(season_header_text[1:]) # Extract season number from header like "S01"
            cell = ws.cell(row=row_idx, column=current_col)
            season_fill, season_text = None, "" # Default values
            plex_s_ep_count = plex_seasons_data.get(season_num_from_header, {}).get('episodes_in_plex', 0)
            
            if tvmaze_info and tvmaze_info.get('seasons'):
                tvmaze_s_data = tvmaze_info['seasons'].get(season_num_from_header)
                if tvmaze_s_data: # If this season exists on TVMaze
                    tvmaze_s_ep_count = tvmaze_s_data.get('total_episodes', 0)
                    season_text = f"{plex_s_ep_count}/{tvmaze_s_ep_count}"
                    # Determine fill color for the season cell.
                    if tvmaze_s_ep_count == 0: season_fill = STYLES['fills']['gray'] if plex_s_ep_count == 0 else STYLES['fills']['yellow']
                    elif plex_s_ep_count >= tvmaze_s_ep_count: season_fill = STYLES['fills']['green']
                    elif plex_s_ep_count > 0: season_fill = STYLES['fills']['red']
                else: # Season not listed on TVMaze
                    if plex_s_ep_count > 0: season_text, season_fill = f"{plex_s_ep_count}/?", STYLES['fills']['yellow']
                    else: season_fill = STYLES['fills']['gray'] # No Plex episodes, not on TVMaze
            else: # No TVMaze info for the show
                if plex_s_ep_count > 0: season_text, season_fill = f"{plex_s_ep_count}/?", STYLES['fills']['yellow']
                # Gray out only if season number is beyond what Plex has any record of.
                elif season_num_from_header > max(plex_seasons_data.keys(), default=-1): season_fill = STYLES['fills']['gray']
            
            cell.value = season_text
            apply_cell_styling(cell, fill_pattern=season_fill, alignment_key='center')
            current_col += 1
            
    # Adjust column widths for the TV show sheet.
    auto_adjust_columns(ws, min_width=7, max_width=15, wrap_text_columns=['Summary', 'Tagline', 'Collections', 'Genres', 'Labels'])
    # Specific widths for Title and Studio if they are selected.
    if 'Title' in SELECTED_SHOW_FIELDS:
        try: ws.column_dimensions[get_column_letter(SELECTED_SHOW_FIELDS.index('Title') + 1)].width = 35
        except ValueError: pass # Should not happen if logic is correct
    if 'Studio' in SELECTED_SHOW_FIELDS:
        try: ws.column_dimensions[get_column_letter(SELECTED_SHOW_FIELDS.index('Studio') + 1)].width = 20
        except ValueError: pass

    # Create an Excel table for the TV show data.
    if len(shows_data_full) > 0 and len(final_headers) > 0:
        last_col_letter = get_column_letter(len(final_headers))
        table_name_base = "".join(c if c.isalnum() else "_" for c in safe_sheet_name)
        create_table(ws, f"{table_name_base}_Table", f"A1:{last_col_letter}{len(shows_data_full) + 1}", style_name='TableStyleMedium9') 

def check_file_writable(filename: str) -> bool:
    """Checks if a file is writable or if its directory is writable for creation."""
    if os.path.exists(filename): return os.access(filename, os.W_OK)
    # If file doesn't exist, check if the parent directory is writable.
    p_dir = os.path.dirname(filename) or '.'; return os.access(p_dir, os.W_OK)

def main():
    """Main function to orchestrate the Plex media export process."""
    print("Plex Media Export Script started.")
    
    # Critical check: Ensure Plex URL and Token are provided.
    if not PLEX_URL or not PLEX_TOKEN: 
        print(f"Critical Error: {PLEX_URL_ENV_VAR} and/or {PLEX_TOKEN_ENV_VAR} are not set in your .env file or environment.")
        sys.exit(1)
        
    # Connect to Plex server.
    plex = connect_to_plex(); 
    if not plex: sys.exit(1) # Exit if connection failed.
    
    # Setup output filename with a timestamp.
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    base_filename = f"PlexMediaExport_{timestamp}.xlsx"
    output_dir = PLEX_EXPORT_DIR # Get output directory from config (defaults to current)
    
    # Ensure output directory exists, create if not.
    if not os.path.isdir(output_dir):
        try: 
            os.makedirs(output_dir, exist_ok=True)
            print(f"Created output directory: {os.path.abspath(output_dir)}")
        except OSError as e: 
            print(f"Error: Could not create output directory {output_dir}: {e}"); sys.exit(1)
            
    filename = os.path.join(output_dir, base_filename) # Full path to the output file.
    
    # Check if the output file is writable.
    if not check_file_writable(filename): 
        print(f"Error: Cannot write to target file {filename}. Check permissions or path."); sys.exit(1)
        
    # Create a new Excel workbook and remove the default sheet.
    wb = Workbook(); 
    if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])
    
    total_movies_processed, total_shows_processed = 0, 0
    
    # Fetch library sections from Plex.
    try:
        print("Fetching library sections from Plex..."); 
        sections = [s for s in plex.library.sections() if s.type in ['movie', 'show']] # Filter for movie or show sections
        print(f"Found {len(sections)} movie/show sections.")
    except Exception as e: 
        print(f"Error fetching library sections from Plex: {e}"); sys.exit(1)
        
    # If no relevant sections found, save an empty report and exit.
    if not sections:
        print("No movie or show sections found in your Plex library.")
        try: 
            wb.save(filename); print(f"Empty report saved as: {os.path.abspath(filename)}")
        except Exception as e: 
            print(f"Error saving empty workbook: {e}")
        sys.exit(0)
        
    # Process each section.
    for section in sections:
        print(f"\nProcessing section: {section.title} (Type: {section.type})")
        try:
            items = section.all() # Get all items (movies or shows) in the current section.
            if not items: 
                print(f"  No items found in section '{section.title}'."); continue
            print(f"  Found {len(items)} items in '{section.title}'.")
            
            # Process based on section type.
            if section.type == 'movie':
                print(f"  Fetching details for {len(items)} movies...")
                movie_list = get_movie_details(items)
                if movie_list: 
                    create_movies_worksheet(section.title, wb, movie_list)
                    total_movies_processed += len(movie_list)
                else: 
                    print(f"  No movie details to write for section '{section.title}'.")
            elif section.type == 'show':
                print(f"  Fetching details for {len(items)} TV shows...")
                shows_data, max_seasons = get_show_details(items)
                if shows_data: 
                    create_tv_shows_worksheet(section.title, wb, shows_data, max_seasons)
                    total_shows_processed += len(shows_data)
                else: 
                    print(f"  No TV show details to write for section '{section.title}'.")
        except Exception as e: 
            print(f"  An error occurred while processing section '{section.title}': {e}")
            # import traceback # Uncomment for full traceback during debugging
            # traceback.print_exc()

    # Save the final Excel workbook.
    try:
        wb.save(filename)
        print(f"\n-----------------------------------------------------------\nExport complete!\nProcessed {total_movies_processed} movies and {total_shows_processed} TV series.\nReport saved as: {os.path.abspath(filename)}\n-----------------------------------------------------------")
    except Exception as e: 
        print(f"Error saving workbook to {filename}: {e}")

# Standard Python idiom: ensure main() is called only when the script is executed directly.
if __name__ == "__main__": 
    main()
# End of script
