#!/usr/bin/env python3
################################################################################
# Plex Media Export Script - Optimized Version
# 
# Purpose: Generate a comprehensive Excel report of Plex Media Server content,
#           including movies and TV shows with TVMaze integration.
#
# Features:
# - Automatically loads environment variables from a .env file
# - Customizable movie & TV show export fields via .env file
# - Optimized field fetching: only processes requested fields from Plex objects
# - Parallel processing for movie data collection
# - Session management for improved API performance
# - Memory optimized Excel generation
# - Cached TVMaze lookups
# - Resolution-based highlighting (4K Green, 1080p White, <1080p Yellow)
# - TV show completion tracking
# - Sortable and filterable Excel headers (using TableStyleMedium9 for all sheets with row stripes)
################################################################################

# Standard library imports
from datetime import datetime  # For timestamping the output file and formatting dates
from typing import Dict, List, Optional, Union, Tuple, Callable, Any, TypedDict # For type hinting, improving code readability and maintainability
import sys  # For system-specific parameters and functions, like exiting the script
from concurrent.futures import ThreadPoolExecutor  # For parallel processing of tasks (like fetching movie details)
from functools import lru_cache, wraps  # For caching results and decorator utilities
import os  # For interacting with the operating system (like accessing environment variables and file paths)
import logging  # For professional logging instead of print statements
from logging.handlers import RotatingFileHandler  # For rotating log files
import time  # For retry delay timing
import pickle  # For persistent caching of TVMaze API results

# Third-party library imports
from dotenv import load_dotenv  # For loading environment variables from a .env file
from plexapi.server import PlexServer  # The main class for interacting with a Plex Media Server
import pandas as pd  # For data manipulation and analysis, used here for creating DataFrames for movies
from openpyxl import Workbook  # For creating and manipulating Excel (.xlsx) files
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment  # For styling cells in Excel
from openpyxl.utils import get_column_letter  # Utility to convert column index to letter (e.g., 1 -> A)
from openpyxl.worksheet.table import Table, TableStyleInfo  # For creating formatted tables within Excel worksheets
from requests import Session  # For making HTTP requests with connection pooling and session persistence
import requests  # For requests.exceptions, to catch specific HTTP request errors

# --- Type Definitions ---
# TypedDict definitions for structured data throughout the script (Python 3.8+)

class TVMazeSeasonInfo(TypedDict):
    """Structure for TVMaze season information."""
    total_episodes: int

class TVMazeShowInfo(TypedDict):
    """Structure for TVMaze show information."""
    total_seasons: int
    seasons: Dict[int, TVMazeSeasonInfo]

class PlexSeasonData(TypedDict):
    """Structure for Plex season data."""
    episodes_in_plex: int
    season_number: int

class CacheEntry(TypedDict):
    """Structure for persistent cache entries."""
    timestamp: datetime
    data: Optional[TVMazeShowInfo]

class CacheFile(TypedDict):
    """Structure for the persistent cache file."""
    version: str
    entries: Dict[str, CacheEntry]
    saved_at: datetime

class ShowData(TypedDict):
    """Structure for processed TV show data."""
    seasons_in_plex: Dict[int, PlexSeasonData]
    tvmaze_info: Optional[TVMazeShowInfo]

class ValidationResult(TypedDict):
    """Structure for environment validation results."""
    errors: List[str]
    warnings: List[str]

# --- Load Environment Variables ---
# This function call loads variables defined in a .env file in the script's directory (or parent directories)
# into environment variables, making them accessible via os.environ.get().
# This is useful for storing sensitive information like API keys or URLs outside the script.
load_dotenv()

# --- Configuration ---
# Define keys for environment variables to ensure consistency and avoid typos when accessing them.
PLEX_URL_ENV_VAR = 'PLEX_URL'
PLEX_TOKEN_ENV_VAR = 'PLEX_TOKEN'
PLEX_EXPORT_DIR_ENV_VAR = 'PLEX_EXPORT_DIR'
PLEX_MOVIE_EXPORT_FIELDS_ENV_VAR = 'PLEX_MOVIE_EXPORT_FIELDS'
PLEX_SHOW_EXPORT_FIELDS_ENV_VAR = 'PLEX_SHOW_EXPORT_FIELDS'

# Retrieve Plex server URL and token from environment variables.
# These are essential for connecting to the Plex server.
PLEX_URL = os.environ.get(PLEX_URL_ENV_VAR)
PLEX_TOKEN = os.environ.get(PLEX_TOKEN_ENV_VAR)
# Retrieve the desired output directory for the Excel file. Defaults to the current directory (".") if not set.
PLEX_EXPORT_DIR = os.environ.get(PLEX_EXPORT_DIR_ENV_VAR, ".")
# Define the base URL for the TVMaze API.
TVMAZE_API = 'https://api.tvmaze.com'

# --- Movie Field Configuration ---
# Defines a comprehensive list of all movie fields that this script is capable of processing.
# This list is used to validate user-defined fields from the .env file.
ALL_POSSIBLE_MOVIE_FIELDS = [
    'Title', 'Year', 'Studio', 'ContentRating',
    'Video Resolution', 'Bitrate (kbps)', 'File Path', 'Container', 'Duration (min)',
    'AddedAt', 'LastViewedAt', 'OriginallyAvailableAt', 'Summary', 'Tagline',
    'AudienceRating', 'Rating', 'Collections', 'Genres', 'Labels',
    'AspectRatio', 'AudioChannels', 'AudioCodec', 'VideoCodec', 'VideoFrameRate',
    'Height', 'Width', 'ViewCount', 'SkipCount'
]
# Defines the default set of movie fields to be exported if the user doesn't specify any.
DEFAULT_MOVIE_FIELDS = [
    'Title', 'Year', 'Studio', 'ContentRating',
    'Video Resolution', 'Bitrate (kbps)', 'File Path', 'Container', 'Duration (min)'
]
# Retrieve the user's preferred movie export fields from the .env file.
user_defined_movie_fields_str = os.environ.get(PLEX_MOVIE_EXPORT_FIELDS_ENV_VAR)
if user_defined_movie_fields_str:
    # If the user provided a string, split it by commas, strip whitespace from each field.
    temp_selected_movie_fields = [field.strip() for field in user_defined_movie_fields_str.split(',') if field.strip()]
    # Filter the user's list against ALL_POSSIBLE_MOVIE_FIELDS to ensure only valid fields are used
    # and to maintain a consistent order if the user's order differs (though this implementation uses user's order if valid).
    SELECTED_MOVIE_FIELDS = [field for field in temp_selected_movie_fields if field in ALL_POSSIBLE_MOVIE_FIELDS]
    if not SELECTED_MOVIE_FIELDS:
        # If the user's string was all invalid fields or just commas, fall back to defaults.
        # Note: logger not available yet, will be initialized after configuration
        SELECTED_MOVIE_FIELDS = DEFAULT_MOVIE_FIELDS[:] # Use a copy of the default list
else:
    # If the environment variable is not set, use the default fields.
    SELECTED_MOVIE_FIELDS = DEFAULT_MOVIE_FIELDS[:] # Use a copy

# --- TV Show Field Configuration ---
# Defines all possible base metadata fields for TV shows that this script can process.
# Season/episode specific details are handled separately.
ALL_POSSIBLE_SHOW_FIELDS = [
    'Title', 'Year', 'Studio', 'ContentRating', 'Summary', 'Tagline',
    'AddedAt', 'LastViewedAt', 'OriginallyAvailableAt',
    'AudienceRating', 'Rating', 'Collections', 'Genres', 'Labels',
    'ViewCount', 'SkipCount'
]
# Defines the default set of TV show base fields to export.
DEFAULT_SHOW_FIELDS = [
    'Title', 'Year', 'Studio', 'ContentRating', 'Summary'
]
# Retrieve user's preferred TV show export fields from .env.
user_defined_show_fields_str = os.environ.get(PLEX_SHOW_EXPORT_FIELDS_ENV_VAR)
if user_defined_show_fields_str:
    # Process the user's string similar to how movie fields are processed.
    temp_selected_show_fields = [field.strip() for field in user_defined_show_fields_str.split(',') if field.strip()]
    SELECTED_SHOW_FIELDS = [field for field in temp_selected_show_fields if field in ALL_POSSIBLE_SHOW_FIELDS]
    if not SELECTED_SHOW_FIELDS:
        # Note: logger not available yet, will be initialized after configuration
        SELECTED_SHOW_FIELDS = DEFAULT_SHOW_FIELDS[:]
else:
    SELECTED_SHOW_FIELDS = DEFAULT_SHOW_FIELDS[:]

# --- API Configuration Constants ---
TVMAZE_REQUEST_TIMEOUT = 10  # seconds for TVMaze API requests
TVMAZE_CACHE_SIZE = 256  # maximum number of cached TVMaze lookups (in-memory)
TVMAZE_MAX_RETRIES = 3  # number of retry attempts for failed API calls
TVMAZE_RETRY_DELAY = 1.0  # seconds between retry attempts

# --- Cache Configuration Constants ---
TVMAZE_CACHE_FILE = os.path.join(PLEX_EXPORT_DIR if PLEX_EXPORT_DIR else '.', '.tvmaze_cache.pkl')  # persistent cache file path
TVMAZE_CACHE_MAX_AGE_DAYS = 30  # cache entries older than this are considered stale

# --- Performance Configuration Constants ---
MAX_WORKER_THREADS = 10  # maximum parallel workers for movie processing
WORKER_THREAD_MULTIPLIER = 4  # additional workers beyond CPU count

# --- Excel Formatting Constants ---
COLUMN_MIN_WIDTH = 10  # minimum column width in Excel
COLUMN_MAX_WIDTH = 50  # maximum column width for auto-sized columns
COLUMN_MAX_WIDTH_WRAPPED = 60  # maximum width for columns with text wrapping
EXCEL_TABLE_STYLE = 'TableStyleMedium9'  # Excel table style to apply

# --- Resolution Categories for Color Coding ---
RESOLUTION_4K = ['4k', 'uhd', '2160p']  # 4K resolution identifiers
RESOLUTION_1080P = ['1080', '1080p']  # 1080p resolution identifiers
RESOLUTION_720P = ['720', '720p']  # 720p resolution identifiers
RESOLUTION_SD = ['sd', '480', '480p', '576', '576p', 'dvd', 'pal', 'ntsc']  # SD resolution identifiers

# --- Logging Setup ---
def setup_logging() -> logging.Logger:
    """
    Setup dual logging to console and rotating file.

    Returns:
        logging.Logger: Configured logger instance
    """
    # Determine log directory (relative to export directory if set, otherwise current dir)
    log_dir = os.path.join(PLEX_EXPORT_DIR if PLEX_EXPORT_DIR else '.', 'logs')
    os.makedirs(log_dir, exist_ok=True)

    log_file = os.path.join(log_dir, f'plex_export_{datetime.now().strftime("%Y%m%d")}.log')

    formatter = logging.Formatter(
        '%(asctime)s - %(levelname)s - %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )

    # File handler (5MB max, keep 5 backups)
    file_handler = RotatingFileHandler(log_file, maxBytes=5*1024*1024, backupCount=5)
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(formatter)

    # Console handler
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(formatter)

    # Root logger
    logger = logging.getLogger()
    logger.setLevel(logging.DEBUG)
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)

    return logger

# --- Environment Validation ---
def validate_environment() -> Tuple[List[str], List[str]]:
    """
    Validate required environment variables and configuration.

    Returns:
        Tuple[List[str], List[str]]: A tuple of (errors, warnings) as lists of strings
    """
    errors: List[str] = []
    warnings: List[str] = []

    # Validate PLEX_URL
    if not PLEX_URL:
        errors.append(f"{PLEX_URL_ENV_VAR} is required")
    elif not PLEX_URL.startswith(('http://', 'https://')):
        errors.append(f"{PLEX_URL_ENV_VAR} must start with http:// or https://")

    # Validate PLEX_TOKEN
    if not PLEX_TOKEN:
        errors.append(f"{PLEX_TOKEN_ENV_VAR} is required")
    elif len(PLEX_TOKEN) < 10:
        warnings.append(f"{PLEX_TOKEN_ENV_VAR} seems too short, verify it's correct")

    # Validate export directory
    if PLEX_EXPORT_DIR and not os.path.isabs(PLEX_EXPORT_DIR):
        warnings.append(f"{PLEX_EXPORT_DIR_ENV_VAR} should be an absolute path for clarity")

    # Validate field selections (warn if defaults were used due to invalid config)
    if user_defined_movie_fields_str and not any(field in ALL_POSSIBLE_MOVIE_FIELDS for field in [f.strip() for f in user_defined_movie_fields_str.split(',') if f.strip()]):
        warnings.append(f"{PLEX_MOVIE_EXPORT_FIELDS_ENV_VAR} contains no valid fields, using defaults")

    if user_defined_show_fields_str and not any(field in ALL_POSSIBLE_SHOW_FIELDS for field in [f.strip() for f in user_defined_show_fields_str.split(',') if f.strip()]):
        warnings.append(f"{PLEX_SHOW_EXPORT_FIELDS_ENV_VAR} contains no valid fields, using defaults")

    return errors, warnings


def retry_on_failure(max_retries: int = TVMAZE_MAX_RETRIES, base_delay: float = TVMAZE_RETRY_DELAY, backoff_factor: float = 2.0) -> Callable:
    """
    Decorator to retry function on failure with exponential backoff.

    Args:
        max_retries: Maximum number of retry attempts (default from TVMAZE_MAX_RETRIES constant)
        base_delay: Initial delay between retries in seconds (default from TVMAZE_RETRY_DELAY constant)
        backoff_factor: Multiplier for delay after each retry (default 2.0)

    Returns:
        Callable: Decorated function that retries on requests.exceptions.RequestException
    """
    def decorator(func: Callable) -> Callable:
        @wraps(func)
        def wrapper(*args: Any, **kwargs: Any) -> Any:
            delay = base_delay
            for attempt in range(max_retries + 1):
                try:
                    return func(*args, **kwargs)
                except requests.exceptions.RequestException as e:
                    if attempt < max_retries:
                        # Use a logger that will be available at runtime
                        logging.warning(f"{func.__name__} failed (attempt {attempt + 1}/{max_retries + 1}): {e}. Retrying in {delay:.1f}s...")
                        time.sleep(delay)
                        delay *= backoff_factor
                    else:
                        logging.error(f"{func.__name__} failed after {max_retries + 1} attempts: {e}")
                        raise
                except Exception as e:
                    logging.error(f"{func.__name__} encountered unexpected error: {e}")
                    raise
            return None
        return wrapper
    return decorator


# --- Persistent Cache Management ---
# Global variable to hold the persistent cache
_persistent_tvmaze_cache: Dict[str, CacheEntry] = {}


def load_tvmaze_cache() -> Dict[str, CacheEntry]:
    """
    Loads the TVMaze cache from disk if it exists and is valid.

    Returns:
        Dict[str, CacheEntry]: The loaded cache dictionary, or an empty dict if no valid cache exists.
    """
    global _persistent_tvmaze_cache

    if not os.path.exists(TVMAZE_CACHE_FILE):
        logging.info("No existing TVMaze cache found. Starting with empty cache.")
        return {}

    try:
        with open(TVMAZE_CACHE_FILE, 'rb') as f:
            cache_data: Any = pickle.load(f)

        # Validate cache structure
        if not isinstance(cache_data, dict) or 'version' not in cache_data or 'entries' not in cache_data:
            logging.warning("Invalid cache file structure. Starting with empty cache.")
            return {}

        # Clean expired entries
        now = datetime.now()
        valid_entries: Dict[str, CacheEntry] = {}
        expired_count: int = 0

        for key, entry in cache_data['entries'].items():
            if 'timestamp' in entry and 'data' in entry:
                age_days = (now - entry['timestamp']).days
                if age_days <= TVMAZE_CACHE_MAX_AGE_DAYS:
                    valid_entries[key] = entry
                else:
                    expired_count += 1

        logging.info(f"Loaded TVMaze cache: {len(valid_entries)} valid entries, {expired_count} expired entries removed.")
        return valid_entries

    except Exception as e:
        logging.error(f"Error loading TVMaze cache: {e}. Starting with empty cache.")
        return {}


def save_tvmaze_cache() -> None:
    """
    Saves the current TVMaze cache to disk.
    """
    global _persistent_tvmaze_cache

    try:
        cache_data: CacheFile = {
            'version': '1.0',
            'entries': _persistent_tvmaze_cache,
            'saved_at': datetime.now()
        }

        # Ensure directory exists (using walrus operator for Python 3.8+)
        if (cache_dir := os.path.dirname(TVMAZE_CACHE_FILE)) and not os.path.exists(cache_dir):
            os.makedirs(cache_dir, exist_ok=True)

        with open(TVMAZE_CACHE_FILE, 'wb') as f:
            pickle.dump(cache_data, f)

        logging.info(f"Saved TVMaze cache with {len(_persistent_tvmaze_cache)} entries.")

    except Exception as e:
        logging.error(f"Error saving TVMaze cache: {e}")


def get_from_cache(key: str) -> Optional[TVMazeShowInfo]:
    """
    Retrieves an entry from the persistent cache if it exists and is valid.

    Args:
        key: The cache key (usually search term or IMDB ID)

    Returns:
        Optional[TVMazeShowInfo]: The cached data, or None if not found or expired
    """
    global _persistent_tvmaze_cache

    if key not in _persistent_tvmaze_cache:
        return None

    entry: CacheEntry = _persistent_tvmaze_cache[key]
    age_days = (datetime.now() - entry['timestamp']).days

    if age_days > TVMAZE_CACHE_MAX_AGE_DAYS:
        # Entry expired, remove it
        del _persistent_tvmaze_cache[key]
        return None

    return entry['data']


def add_to_cache(key: str, data: TVMazeShowInfo) -> None:
    """
    Adds an entry to the persistent cache.

    Args:
        key: The cache key (usually search term or IMDB ID)
        data: The data to cache
    """
    global _persistent_tvmaze_cache

    _persistent_tvmaze_cache[key] = {
        'timestamp': datetime.now(),
        'data': data
    }


# --- Initialization ---
# Setup logging first so it's available throughout the script
logger = setup_logging()

# Validate environment configuration
errors, warnings = validate_environment()

# Log warnings
for warning in warnings:
    logger.warning(warning)

# If there are errors, log them and exit
if errors:
    for error in errors:
        logger.error(error)
    logger.error("Configuration validation failed. Please check your .env file.")
    sys.exit(1)

# Log configuration
logger.info(f"Selected movie fields for export: {', '.join(SELECTED_MOVIE_FIELDS)}")
logger.info(f"Selected TV show fields for export: {', '.join(SELECTED_SHOW_FIELDS)}")

# Create a global requests.Session object. This allows for connection pooling,
# which can improve performance when making multiple HTTP requests to the same host (TVMaze API).
session = Session()

# Load persistent TVMaze cache from disk
_persistent_tvmaze_cache = load_tvmaze_cache()

# Define a dictionary of styles to be used for formatting Excel cells.
# This centralizes style definitions for consistency and easier modification.
STYLES = {
    'fills': {  # Define various fill patterns (background colors)
        'gray': PatternFill(patternType='solid', fgColor='D3D3D3'),  # For non-existent TV seasons
        'green': PatternFill(patternType='solid', fgColor='90EE90'), # For complete TV series/seasons
        'red': PatternFill(patternType='solid', fgColor='FFB6B6'),   # For incomplete TV series/seasons
        'yellow': PatternFill(patternType='solid', fgColor='FFFFCC'),# For low-resolution movies (<1080p)
        '4k_specific_light_green': PatternFill(patternType='solid', fgColor='77b190'), # For 4K movies (dark green)
        'white': PatternFill(patternType='solid', fgColor='FFFFFF')  # Explicit white fill (for 1080p movies)
    },
    'borders': { # Define border styles
        'thin': Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')),
        'header': Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thick')) # Thicker bottom border for headers
    },
    'fonts': { # Define font styles
        'bold': Font(bold=True) # Bold font for headers
    },
    'alignments': { # Define text alignment styles
        'center': Alignment(horizontal='center', vertical='center', wrap_text=False), # Center alignment, no text wrapping
        'left': Alignment(horizontal='left', vertical='center', wrap_text=False),   # Left alignment, no text wrapping
        'summary': Alignment(horizontal='left', vertical='top', wrap_text=True)     # Left-top alignment with text wrapping (for summaries)
    }
}

def connect_to_plex() -> Optional[PlexServer]:
    """
    Establishes a connection to the Plex Media Server using credentials from environment variables.

    Returns:
        Optional[PlexServer]: A PlexServer instance if connection is successful, None otherwise.
    """
    # Check if PLEX_URL and PLEX_TOKEN are set.
    if not PLEX_URL or not PLEX_TOKEN:
        logger.error(f"Plex URL ({PLEX_URL_ENV_VAR}) and Token ({PLEX_TOKEN_ENV_VAR}) must be set.")
        return None
    try:
        # Create a new session for Plex API calls. This is separate from the global session for TVMaze.
        plex_session = Session()
        # Note: SSL verification can be disabled here if needed (e.g., for self-signed certs on a local network),
        # but it's generally not recommended for security reasons.
        # Example:
        # from urllib3.exceptions import InsecureRequestWarning
        # requests.packages.urllib3.disable_warnings(category=InsecureRequestWarning)
        # plex_session.verify = False

        logger.info(f"Connecting to Plex server at {PLEX_URL}...")
        # Initialize the PlexServer object with the URL, token, and session.
        server = PlexServer(PLEX_URL, PLEX_TOKEN, session=plex_session)
        # A simple test to confirm connection by fetching the server's friendly name.
        logger.info(f"Successfully connected to Plex server: {server.friendlyName}")
        return server
    except Exception as e:
        # Catch any exception during connection and log an error message.
        logger.error(f"Failed to connect to Plex server: {e}")
        return None

@retry_on_failure() # Decorator to retry on network failures with exponential backoff
def _fetch_tvmaze_show_info_from_api(search_term: str, is_imdb_id: bool = False) -> Optional[TVMazeShowInfo]:
    """
    Internal function that fetches TV show information directly from the TVMaze API.
    This is called only when data is not found in persistent or in-memory cache.

    Args:
        search_term: The name of the TV show or its IMDB ID.
        is_imdb_id: True if search_term is an IMDB ID, False if it's a show title.

    Returns:
        Optional[TVMazeShowInfo]: A dictionary containing total seasons and episode counts per season,
                                  or None if the show is not found or an error occurs.
    """
    try:
        # Determine the TVMaze API endpoint and parameters based on whether searching by IMDB ID or title.
        api_url, params = ((f"{TVMAZE_API}/lookup/shows", {'imdb': search_term})
                           if is_imdb_id
                           else (f"{TVMAZE_API}/search/shows", {'q': search_term}))

        # Make the GET request to TVMaze API using the global session.
        response = session.get(api_url, params=params, timeout=TVMAZE_REQUEST_TIMEOUT)
        response.raise_for_status() # Raise an HTTPError for bad status codes (4xx or 5xx)

        json_response = response.json() # Parse the JSON response

        # Extract the show ID from the response (using walrus operator for Python 3.8+)
        show_id = None
        if is_imdb_id and json_response: # /lookup/shows returns a single show object or null
            show_id = json_response.get('id')
        elif json_response and isinstance(json_response, list) and json_response: # /search/shows returns a list
            # Assume the first result is the most relevant one.
            show_id = json_response[0].get('show', {}).get('id')

        # If a show ID was found, fetch episode details.
        if show_id:
            episodes_response = session.get(f"{TVMAZE_API}/shows/{show_id}/episodes", timeout=TVMAZE_REQUEST_TIMEOUT)
            episodes_response.raise_for_status()
            episodes = episodes_response.json()
            
            seasons_data = {}
            # If no episodes are listed, return 0 seasons.
            if not episodes:
                return {'total_seasons': 0, 'seasons': {}}
            
            # Aggregate episode counts by season number.
            for episode in episodes:
                season_num = episode.get('season')
                if season_num is None: # Skip episodes with no season number (e.g., some specials)
                    continue
                seasons_data.setdefault(season_num, {'total_episodes': 0})['total_episodes'] += 1
            
            # If no valid season data was processed (e.g., all episodes lacked season numbers).
            if not seasons_data:
                return {'total_seasons': 0, 'seasons': {}}

            # Return the total number of seasons and the per-season episode counts.
            return {'total_seasons': max(seasons_data.keys(), default=0), 'seasons': seasons_data}
            
    # Handle specific exceptions that might occur during API requests or JSON parsing.
    except requests.exceptions.RequestException as e:
        logger.error(f"TVMaze API request error for '{search_term}': {e}")
    except ValueError as e: # Includes JSONDecodeError
        logger.error(f"TVMaze API JSON decode error for '{search_term}': {e}")
    except Exception as e: # Catch any other unexpected errors
        logger.error(f"Unexpected error fetching TVMaze data for '{search_term}': {e}")
    return None # Return None if any error occurs


def get_tvmaze_show_info(search_term: str, is_imdb_id: bool = False) -> Optional[TVMazeShowInfo]:
    """
    Fetches TV show information with persistent caching support.

    This function first checks the persistent cache, then falls back to the API if needed.
    Results from the API are automatically added to the persistent cache.

    Args:
        search_term: The name of the TV show or its IMDB ID.
        is_imdb_id: True if search_term is an IMDB ID, False if it's a show title.

    Returns:
        Optional[TVMazeShowInfo]: A dictionary containing total seasons and episode counts per season,
                                  or None if the show is not found or an error occurs.
    """
    # Create cache key (prefix with type for clarity)
    cache_key = f"imdb:{search_term}" if is_imdb_id else f"title:{search_term}"

    # Try to get from persistent cache first
    cached_result = get_from_cache(cache_key)
    if cached_result is not None:
        logger.debug(f"Cache hit for TVMaze lookup: {search_term}")
        return cached_result

    # Cache miss - fetch from API
    logger.debug(f"Cache miss for TVMaze lookup: {search_term}, fetching from API")
    result = _fetch_tvmaze_show_info_from_api(search_term, is_imdb_id)

    # Add to cache if we got a valid result (including explicit None for "not found")
    if result is not None:
        add_to_cache(cache_key, result)

    return result


def format_plex_datetime(dt_obj: Optional[Union[datetime, pd.Timestamp, Any]]) -> str:
    """
    Formats a datetime object (often from Plex attributes) into an ISO 8601 string.
    Handles cases where the datetime object might be None.

    Args:
        dt_obj: The datetime object to format.

    Returns:
        str: The formatted datetime string or 'N/A' if input is None.
    """
    if dt_obj is None:
        return 'N/A'
    try:
        # Check if it's a datetime or pandas Timestamp object before calling isoformat.
        return dt_obj.isoformat() if isinstance(dt_obj, (datetime, pd.Timestamp)) else str(dt_obj)
    except AttributeError: # If it's not a datetime object but something else that doesn't have isoformat
        return str(dt_obj)

def format_plex_list(plex_list_attr: Optional[List[Any]]) -> str:
    """
    Formats a list of Plex Tag objects (e.g., genres, collections, labels)
    into a comma-separated string of their 'tag' or 'title' attributes.

    Args:
        plex_list_attr: A list of Plex Tag-like objects.

    Returns:
        str: A comma-separated string of tags/titles, or 'N/A' if the list is empty/None.
    """
    if plex_list_attr is None or not plex_list_attr: # Check if None or empty list
        return 'N/A'
    try:
        # Plex Tag objects usually have a 'tag' attribute (for Genre, Director, etc.)
        # or a 'title' attribute (for Collection). This attempts to get the most relevant one.
        return ', '.join(
            item.tag if hasattr(item, 'tag') else
            item.title if hasattr(item, 'title') else
            str(item) # Fallback to string representation if neither 'tag' nor 'title' exists
            for item in plex_list_attr
        )
    except Exception: # Catch any error during list processing
        return 'Error processing list'


# --- Field Processor Dictionaries ---
# Dictionary-based approach for field extraction, eliminating long if-elif chains.
# Each field maps to a callable that extracts the value from the Plex object.

def _get_movie_field_processors() -> Dict[str, Callable[[Any, Any, Any], Union[str, int, float]]]:
    """
    Returns a dictionary of field processors for movie fields.
    Each processor is a callable that takes (movie, media, parts) and returns the field value.

    Returns:
        Dict[str, Callable]: A mapping of field names to processor functions
    """
    return {
        'Title': lambda m, _, __: getattr(m, 'title', 'N/A'),
        'Year': lambda m, _, __: getattr(m, 'year', 'N/A'),
        'Studio': lambda m, _, __: getattr(m, 'studio', None) or 'N/A',
        'ContentRating': lambda m, _, __: getattr(m, 'contentRating', None) or 'N/A',
        'Duration (min)': lambda m, _, __: round(m.duration / 60000) if hasattr(m, 'duration') and m.duration else 'N/A',
        'AddedAt': lambda m, _, __: format_plex_datetime(m.addedAt) if hasattr(m, 'addedAt') else 'N/A',
        'LastViewedAt': lambda m, _, __: format_plex_datetime(m.lastViewedAt) if hasattr(m, 'lastViewedAt') else 'N/A',
        'OriginallyAvailableAt': lambda m, _, __: format_plex_datetime(m.originallyAvailableAt) if hasattr(m, 'originallyAvailableAt') else 'N/A',
        'Summary': lambda m, _, __: getattr(m, 'summary', None) or 'N/A',
        'Tagline': lambda m, _, __: getattr(m, 'tagline', None) or 'N/A',
        'AudienceRating': lambda m, _, __: getattr(m, 'audienceRating', None) or 'N/A',
        'Rating': lambda m, _, __: getattr(m, 'rating', None) or 'N/A',
        'Collections': lambda m, _, __: format_plex_list(m.collections) if hasattr(m, 'collections') else 'N/A',
        'Genres': lambda m, _, __: format_plex_list(m.genres) if hasattr(m, 'genres') else 'N/A',
        'Labels': lambda m, _, __: format_plex_list(m.labels) if hasattr(m, 'labels') else 'N/A',
        'ViewCount': lambda m, _, __: getattr(m, 'viewCount', 0) or 0,
        'SkipCount': lambda m, _, __: getattr(m, 'skipCount', 0) or 0,
        # Media-dependent fields
        'Video Resolution': lambda _, media, __: str(getattr(media, 'videoResolution', '')).strip() if media and getattr(media, 'videoResolution', '').strip() else 'Unknown',
        'Bitrate (kbps)': lambda _, media, __: getattr(media, 'bitrate', None) if media and getattr(media, 'bitrate', None) else 'Unknown',
        'Container': lambda _, media, __: getattr(media, 'container', None) if media and getattr(media, 'container', None) else 'Unknown',
        'AspectRatio': lambda _, media, __: getattr(media, 'aspectRatio', None) or 'N/A' if media else 'N/A',
        'AudioChannels': lambda _, media, __: getattr(media, 'audioChannels', None) or 'N/A' if media else 'N/A',
        'AudioCodec': lambda _, media, __: getattr(media, 'audioCodec', None) or 'N/A' if media else 'N/A',
        'VideoCodec': lambda _, media, __: getattr(media, 'videoCodec', None) or 'N/A' if media else 'N/A',
        'VideoFrameRate': lambda _, media, __: getattr(media, 'videoFrameRate', None) or 'N/A' if media else 'N/A',
        'Height': lambda _, media, __: getattr(media, 'height', None) or 'N/A' if media else 'N/A',
        'Width': lambda _, media, __: getattr(media, 'width', None) or 'N/A' if media else 'N/A',
        'File Path': lambda _, __, parts: parts.file if parts and hasattr(parts, 'file') else 'Unknown',
    }


def _get_show_field_processors() -> Dict[str, Callable[[Any], Union[str, int, float]]]:
    """
    Returns a dictionary of field processors for TV show fields.
    Each processor is a callable that takes (show_obj) and returns the field value.

    Returns:
        Dict[str, Callable]: A mapping of field names to processor functions
    """
    return {
        'Title': lambda s: getattr(s, 'title', 'N/A'),
        'Year': lambda s: getattr(s, 'year', 'N/A'),
        'Studio': lambda s: getattr(s, 'studio', None) or 'N/A',
        'ContentRating': lambda s: getattr(s, 'contentRating', None) or 'N/A',
        'Summary': lambda s: getattr(s, 'summary', None) or 'N/A',
        'Tagline': lambda s: getattr(s, 'tagline', None) or 'N/A',
        'AddedAt': lambda s: format_plex_datetime(s.addedAt) if hasattr(s, 'addedAt') else 'N/A',
        'LastViewedAt': lambda s: format_plex_datetime(s.lastViewedAt) if hasattr(s, 'lastViewedAt') else 'N/A',
        'OriginallyAvailableAt': lambda s: format_plex_datetime(s.originallyAvailableAt) if hasattr(s, 'originallyAvailableAt') else 'N/A',
        'AudienceRating': lambda s: getattr(s, 'audienceRating', None) or 'N/A',
        'Rating': lambda s: getattr(s, 'rating', None) or 'N/A',
        'Collections': lambda s: format_plex_list(s.collections) if hasattr(s, 'collections') else 'N/A',
        'Genres': lambda s: format_plex_list(s.genres) if hasattr(s, 'genres') else 'N/A',
        'Labels': lambda s: format_plex_list(s.labels) if hasattr(s, 'labels') else 'N/A',
        'ViewCount': lambda s: getattr(s, 'viewCount', 0) or 0,
        'SkipCount': lambda s: getattr(s, 'skipCount', 0) or 0,
    }


def process_movie(movie: Any) -> Dict[str, Union[str, int, float]]:
    """
    Processes a single Plex movie object and extracts data for the fields
    specified in SELECTED_MOVIE_FIELDS using a dictionary-based approach.

    Args:
        movie: A Plex movie object.

    Returns:
        Dict[str, Union[str, int, float]]: A dictionary containing the extracted movie data.
    """
    movie_data = {}

    # Fetch media and parts objects once if any media-dependent fields are requested
    media_dependent_fields = {
        'Video Resolution', 'Bitrate (kbps)', 'File Path', 'Container',
        'AspectRatio', 'AudioChannels', 'AudioCodec', 'VideoCodec',
        'VideoFrameRate', 'Height', 'Width'
    }

    media = None
    parts = None
    if any(field in SELECTED_MOVIE_FIELDS for field in media_dependent_fields):
        media = movie.media[0] if movie.media else None
        if media:
            parts = media.parts[0] if media.parts else None

    # Get field processors dictionary
    field_processors = _get_movie_field_processors()

    # Process each selected field using the dictionary lookup
    for field_name in SELECTED_MOVIE_FIELDS:
        try:
            processor = field_processors.get(field_name)
            if processor:
                value = processor(movie, media, parts)
            else:
                logger.warning(f"Unknown field '{field_name}' requested for movie export")
                value = 'N/A'
        except Exception as e:
            logger.error(f"Error processing field '{field_name}' for movie '{getattr(movie, 'title', 'Unknown Title')}': {e}")
            value = 'Error Processing Field'

        movie_data[field_name] = value

    return movie_data

def get_movie_details(movies: List[Any]) -> List[Dict[str, Union[str, int, float]]]:
    """
    Retrieves details for a list of Plex movie objects in parallel using a ThreadPoolExecutor.

    Args:
        movies: A list of Plex movie objects.

    Returns:
        List[Dict[str, Union[str, int, float]]]: A list of dictionaries, each containing movie details.
    """
    # Determine a reasonable number of worker threads.
    num_workers = min(MAX_WORKER_THREADS, (os.cpu_count() or 1) + WORKER_THREAD_MULTIPLIER)
    # Use ThreadPoolExecutor to process movies in parallel.
    # executor.map applies process_movie to each movie in the list.
    with ThreadPoolExecutor(max_workers=num_workers) as executor:
        # filter(None, ...) removes any None results if process_movie were to return None (it currently doesn't).
        return list(filter(None, executor.map(process_movie, movies)))

def process_show_metadata(show_obj: Any) -> Dict[str, Union[str, int, float]]:
    """
    Processes a single Plex TV show object and extracts base metadata for the fields
    specified in SELECTED_SHOW_FIELDS using a dictionary-based approach.

    Args:
        show_obj: A Plex TV show object.

    Returns:
        Dict[str, Union[str, int, float]]: A dictionary containing the extracted TV show metadata.
    """
    show_metadata = {}

    # Get field processors dictionary
    field_processors = _get_show_field_processors()

    # Process each selected field using the dictionary lookup
    for field_name in SELECTED_SHOW_FIELDS:
        try:
            processor = field_processors.get(field_name)
            if processor:
                value = processor(show_obj)
            else:
                logger.warning(f"Unknown field '{field_name}' requested for TV show export")
                value = 'N/A'
        except Exception as e:
            logger.error(f"Error processing field '{field_name}' for show '{getattr(show_obj, 'title', 'Unknown Title')}': {e}")
            value = 'Error Processing Field'

        show_metadata[field_name] = value

    return show_metadata

def process_single_show(show_obj: Any) -> Dict[str, Any]:
    """
    Processes a single TV show to retrieve its metadata, TVMaze info, and Plex season data.

    This function is designed to be called in parallel by ThreadPoolExecutor.

    Args:
        show_obj: A Plex TV show object.

    Returns:
        Dict[str, Any]: A dictionary containing show metadata, Plex season data, TVMaze info, and max seasons.
    """
    try:
        logger.info(f"Processing TV Show: {show_obj.title}")

        # Get the base metadata for the show using the selected fields.
        show_metadata = process_show_metadata(show_obj)

        # Attempt to find IMDB ID for TVMaze lookup (using walrus operator for Python 3.8+)
        imdb_id = (imdb_id_full.split('imdb://')[-1]
                   if (imdb_id_full := next((guid.id for guid in show_obj.guids if guid.id.startswith('imdb://')), None))
                   else None)

        # Optimized TVMaze lookup with fallback logic
        tvmaze_info = None
        if imdb_id and (tvmaze_info := get_tvmaze_show_info(imdb_id, is_imdb_id=True)):
            logger.debug(f"  TVMaze lookup successful via IMDB ID: {imdb_id}")

        # Fallback to title search only if IMDB lookup failed and we have a valid title (using walrus operator)
        if not tvmaze_info and (search_title := show_obj.originalTitle or show_obj.title) and search_title.strip():
            logger.debug(f"  Attempting TVMaze lookup by title: {search_title}")
            if (tvmaze_info := get_tvmaze_show_info(search_title)):
                logger.debug(f"  TVMaze lookup successful via title search")

        # Track max seasons for this show
        show_max_seasons = 0
        if tvmaze_info and tvmaze_info.get('total_seasons') is not None:
            show_max_seasons = tvmaze_info['total_seasons']
        elif not tvmaze_info:
            logger.warning(f"  Could not find TVMaze info for: {show_obj.title} (IMDB: {imdb_id or 'N/A'})")

        # Get season and episode counts from Plex.
        plex_seasons_data = {}
        try:
            for s in show_obj.seasons():
                if s.seasonNumber is not None:  # Only process seasons with a valid number
                    try:
                        season_num_int = int(s.seasonNumber)
                        plex_seasons_data[season_num_int] = {
                            'episodes_in_plex': s.leafCount,  # Use leafCount for O(1) access instead of len(s.episodes())
                            'season_number': season_num_int
                        }
                    except ValueError:  # Handle cases where seasonNumber might not be a valid integer
                        logger.warning(f"  Skipping invalid season number '{s.seasonNumber}' for show '{show_obj.title}'")
        except Exception as e:
            logger.error(f"  Error fetching Plex season details for {show_obj.title}: {e}")

        # Combine the base metadata, Plex season data, TVMaze info, and max seasons
        combined_show_data = {
            **show_metadata,
            'seasons_in_plex': plex_seasons_data,
            'tvmaze_info': tvmaze_info,
            'max_seasons': show_max_seasons
        }

        return combined_show_data

    except Exception as e:
        logger.error(f"Error processing show {show_obj.title}: {e}")
        # Return minimal data on error
        return {
            'Title': show_obj.title,
            'seasons_in_plex': {},
            'tvmaze_info': None,
            'max_seasons': 0
        }


def get_show_details(shows: List[Any]) -> Tuple[List[Dict[str, Any]], int]:
    """
    Retrieves detailed information for a list of TV shows using parallel processing.

    This function uses ThreadPoolExecutor to process multiple shows concurrently,
    significantly improving performance for large libraries.

    Args:
        shows: A list of Plex TV show objects.

    Returns:
        Tuple[List[Dict[str, Any]], int]: A tuple containing:
            - A list of dictionaries, each with details for a TV show.
            - An integer representing the maximum number of seasons found across all shows.
    """
    processed_shows_data = []
    max_seasons_overall = 0

    # Calculate optimal number of workers
    num_workers = min(MAX_WORKER_THREADS, len(shows))

    logger.info(f"Processing {len(shows)} TV shows with {num_workers} parallel workers...")

    # Use ThreadPoolExecutor for parallel processing
    with ThreadPoolExecutor(max_workers=num_workers) as executor:
        # Submit all show processing tasks and get results as they complete
        show_results = list(executor.map(process_single_show, shows))

    # Process results and find max seasons
    for show_data in show_results:
        if show_data:
            show_max = show_data.pop('max_seasons', 0)  # Remove and get max_seasons from the dict
            max_seasons_overall = max(max_seasons_overall, show_max)
            processed_shows_data.append(show_data)

    logger.info(f"Completed processing {len(processed_shows_data)} TV shows.")

    return processed_shows_data, max_seasons_overall

def apply_cell_styling(cell: Any, is_header: bool = False, alignment_key: str = 'center', fill_pattern: Optional[PatternFill] = None) -> None:
    """
    Applies predefined border, alignment, font (if header), and fill styles to an Excel cell.

    Args:
        cell: The openpyxl cell object to style.
        is_header: True if the cell is a header cell, False otherwise.
        alignment_key: The key for the desired alignment style from STYLES['alignments'].
        fill_pattern: The PatternFill object to apply as background, or None for no fill.
    """
    # Apply border style (different for header vs. data cells).
    cell.border = STYLES['borders']['header' if is_header else 'thin']
    # Apply alignment style, defaulting to 'left' if the key is not found.
    alignment_style = STYLES['alignments'].get(alignment_key, STYLES['alignments']['left'])
    cell.alignment = alignment_style
    # Apply bold font if it's a header cell.
    if is_header:
        cell.font = STYLES['fonts']['bold']
    # Apply fill pattern if one is provided.
    if fill_pattern:
        cell.fill = fill_pattern

def create_table(ws: Any, table_name: str, data_range: str, style_name: str = EXCEL_TABLE_STYLE) -> Optional[Table]:
    """
    Creates an Excel table on the given worksheet with specified name, range, and style.
    Ensures table names are sanitized (no spaces) and enables row stripes by default.

    Args:
        ws: The openpyxl worksheet object.
        table_name: The desired name for the table (will be sanitized).
        data_range: The cell range for the table (e.g., "A1:G100").
        style_name: The name of the built-in Excel table style to apply.

    Returns:
        Optional[Table]: The created openpyxl Table object, or None if an error occurs.
    """
    # Sanitize table name: Excel table names cannot contain spaces or certain characters,
    # and must not start with a number.
    sanitized_table_name = table_name.replace(" ", "_").replace("-", "_")
    if not sanitized_table_name or not sanitized_table_name[0].isalpha():
        sanitized_table_name = "T_" + sanitized_table_name # Prepend "T_" if name is invalid
        
    try:
        # Create the Table object.
        table = Table(displayName=sanitized_table_name, ref=data_range)
        # Define the table style information.
        # showRowStripes=True enables alternating row colors from the chosen style.
        # Set to False if you want custom cell fills to always take precedence without interference.
        table.tableStyleInfo = TableStyleInfo(
            name=style_name, showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False
        )
        # Add the table to the worksheet.
        ws.add_table(table)
        return table
    except Exception as e:
        logger.error(f"Error creating table '{sanitized_table_name}': {e}")
        return None

def auto_adjust_columns(ws: Any, min_width: int = COLUMN_MIN_WIDTH, max_width: int = COLUMN_MAX_WIDTH, wrap_text_columns: Optional[List[str]] = None) -> None:
    """
    Automatically adjusts column widths on a worksheet based on content length.
    Allows specifying columns that should rely on text wrapping instead of auto-sizing.

    Args:
        ws: The openpyxl worksheet object.
        min_width: The minimum width for any column.
        max_width: The maximum width for auto-sized columns.
        wrap_text_columns: A list of header names for columns that should be set to max_width and rely on wrap text.
    """
    if wrap_text_columns is None: # Initialize to empty list if not provided
        wrap_text_columns = []
        
    header_row = [cell.value for cell in ws[1]] # Get header values from the first row of the worksheet

    # Iterate through each column in the worksheet.
    for col_idx, column_cells in enumerate(ws.columns):
        column_letter = column_cells[0].column_letter # Get the column letter (e.g., 'A', 'B')
        current_header = header_row[col_idx] if col_idx < len(header_row) else ""

        # If the current column is designated for text wrapping, set its width to max_width and skip auto-sizing.
        if current_header in wrap_text_columns:
            ws.column_dimensions[column_letter].width = max_width 
            continue
            
        try:
            # Calculate the maximum length of content in the current column.
            base_length = max(len(str(cell.value)) for cell in column_cells if cell.value is not None)
            length = base_length + 2 # Add a little padding
            # Constrain the width between min_width and max_width.
            final_width = max(min_width, min(length, max_width))
            ws.column_dimensions[column_letter].width = final_width
        except ValueError: # Handles case where column might be entirely empty, max() on empty sequence
             ws.column_dimensions[column_letter].width = min_width
        except Exception as e: # Catch any other errors during column adjustment
            logger.error(f"Error auto-adjusting column {column_letter}: {e}. Setting default width.")
            ws.column_dimensions[column_letter].width = min_width

def create_movies_worksheet(section_name: str, wb: Workbook, movie_list: List[Dict[str, Union[str, int, float]]]) -> None:
    """
    Creates and populates a worksheet for movies.

    Args:
        section_name: The name of the Plex library section (used for sheet name).
        wb: The openpyxl Workbook object.
        movie_list: A list of dictionaries, each containing movie data.
    """
    if not movie_list: # If no movies were found/processed for this section
        logger.info(f"No movies found in section '{section_name}'. Skipping sheet creation.")
        return

    # Sanitize the section name to be a valid Excel sheet name (max 31 chars, no invalid chars).
    safe_sheet_name = "".join(c if c.isalnum() or c in " _-" else "_" for c in section_name[:31])
    ws = wb.create_sheet(safe_sheet_name) # Create the new worksheet
    ws.freeze_panes = 'A2' # Freeze the header row (row 1) and first column (A) is not frozen here.

    # The movie_list already contains dictionaries with only SELECTED_MOVIE_FIELDS due to process_movie.
    df_data = movie_list
    if not df_data: # Should not happen if movie_list is not empty, but as a safeguard.
        logger.warning(f"No data to process for movies in section '{section_name}' after field selection.")
        return

    # Create a Pandas DataFrame from the movie data, ensuring columns are in the order of SELECTED_MOVIE_FIELDS.
    df = pd.DataFrame(df_data, columns=SELECTED_MOVIE_FIELDS)
    if df.empty: # If DataFrame is empty after creation (e.g., if df_data was empty)
        logger.warning(f"DataFrame is empty for section '{section_name}' after field selection. Skipping sheet content.")
        return
    
    # Determine the field to sort by: 'Title' if selected, otherwise the first selected field.
    sort_by_field = 'Title' if 'Title' in SELECTED_MOVIE_FIELDS else (SELECTED_MOVIE_FIELDS[0] if SELECTED_MOVIE_FIELDS else None)
    if sort_by_field:
        df = df.sort_values(by=sort_by_field, ascending=True) # Sort the DataFrame
    
    headers = SELECTED_MOVIE_FIELDS # The headers are the selected fields

    # Write and style the header row.
    for col_idx, header_text in enumerate(headers, 1):
        apply_cell_styling(ws.cell(row=1, column=col_idx, value=header_text), is_header=True, alignment_key='center')

    # Pre-compute column indices for O(1) lookup instead of O(n) in loops
    resolution_col_idx = SELECTED_MOVIE_FIELDS.index('Video Resolution') if 'Video Resolution' in SELECTED_MOVIE_FIELDS else None

    # --- Debugging Movie Fills ---
    # This section was previously used for debugging and can be re-enabled if needed.
    # print(f"\n--- Debugging Movie Fills for Section: {section_name} ---")

    # Iterate through each row of movie data in the DataFrame.
    for row_idx, row_data_tuple in enumerate(df.itertuples(index=False), 2): # Start from Excel row 2
        row_fill = None # Default to no specific fill for the row.

        # Determine row fill color based on 'Video Resolution' if that field is selected for export.
        if resolution_col_idx is not None:
            try:
                # Access the 'Video Resolution' value using pre-computed index
                raw_resolution_from_tuple = row_data_tuple[resolution_col_idx]
                
                resolution_val = str(raw_resolution_from_tuple).strip().lower() # Process for comparison

                # Apply fill based on resolution.
                if resolution_val in RESOLUTION_4K: row_fill = STYLES['fills']['4k_specific_light_green']
                elif resolution_val in RESOLUTION_1080P: row_fill = STYLES['fills']['white']
                elif resolution_val in RESOLUTION_720P + RESOLUTION_SD:
                    row_fill = STYLES['fills']['yellow']
            except (ValueError, IndexError, AttributeError) as e: 
                # Error accessing or processing resolution, so no specific fill.
                # print(f"  DEBUG: Error determining fill for row {row_idx-1}: {e}") # Optional debug
                pass 
        
        # --- Debug Print (can be re-enabled) ---
        # movie_title_debug = getattr(row_data_tuple, 'Title', 'N/A') if 'Title' in SELECTED_MOVIE_FIELDS else 'N/A'
        # print(f"Movie: {movie_title_debug}, ProcessedRes='{resolution_val if 'Video Resolution' in SELECTED_MOVIE_FIELDS else 'N/A'}', DeterminedFill='{row_fill}'")
        # if row_fill:
        #     print(f"  Fill Type: {type(row_fill)}, FG Color: {row_fill.fgColor.rgb if hasattr(row_fill, 'fgColor') and row_fill.fgColor else 'N/A'}")
        
        # Write data for each cell in the current row and apply styles.
        for col_idx, value in enumerate(row_data_tuple, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            # Ensure value is appropriate for Excel cell (numeric or string).
            cell.value = value if isinstance(value, (int, float)) and not pd.isna(value) else str(value if not pd.isna(value) else '')
            
            current_header = headers[col_idx-1] # Get current header name
            # Determine alignment based on header.
            alignment_key = 'summary' if current_header in ['Summary', 'Tagline'] else \
                            'left' if current_header in ['Title', 'File Path', 'Studio', 'Collections', 'Genres', 'Labels'] else 'center'
            apply_cell_styling(cell, alignment_key=alignment_key, fill_pattern=row_fill) # Apply determined row fill
            
    # Create an Excel table for the movie data.
    last_row, last_col_letter = len(df) + 1, get_column_letter(len(headers))
    if last_row > 1 and len(headers) > 0:
        create_table(ws, f'{"".join(c if c.isalnum() else "_" for c in safe_sheet_name)}_Table', f"A1:{last_col_letter}{last_row}")
    # Adjust column widths.
    auto_adjust_columns(ws, max_width=COLUMN_MAX_WIDTH_WRAPPED, wrap_text_columns=['Summary', 'Tagline', 'File Path', 'Collections', 'Genres', 'Labels'])

def _generate_tv_show_headers(shows_data_full: List[Dict[str, Any]], max_seasons_overall: int) -> Tuple[List[str], List[str]]:
    """
    Generates headers for TV show worksheet including base fields, completion status, and season columns.

    Args:
        shows_data_full: Full show data to check for specials.
        max_seasons_overall: Maximum number of seasons across all shows.

    Returns:
        Tuple[List[str], List[str]]: A tuple of (final_headers, season_headers_list)
    """
    base_headers = SELECTED_SHOW_FIELDS[:]
    completion_header = "Series Complete (Plex/TVMaze)"
    has_specials = any(0 in show_dict.get('seasons_in_plex', {}) for show_dict in shows_data_full)
    season_headers_list: List[str] = []
    if has_specials:
        season_headers_list.append("S00")
    season_headers_list.extend([f"S{i:02d}" for i in range(1, max_seasons_overall + 1)])
    final_headers = base_headers + [completion_header] + season_headers_list
    return final_headers, season_headers_list


def _calculate_series_completion(tvmaze_info: Optional[TVMazeShowInfo], plex_seasons_data: Dict[int, PlexSeasonData]) -> Tuple[str, Optional[PatternFill]]:
    """
    Calculates series completion status text and fill color.

    Args:
        tvmaze_info: TVMaze show information.
        plex_seasons_data: Plex season data.

    Returns:
        Tuple[str, Optional[PatternFill]]: A tuple of (status_text, fill_pattern)
    """
    if tvmaze_info and tvmaze_info.get('total_seasons') is not None:
        tvmaze_total_regular_seasons = sum(1 for s_num in tvmaze_info.get('seasons', {}) if s_num > 0)
        complete_plex_regular_seasons_count = 0
        for s_num, tvmaze_s_data in tvmaze_info.get('seasons', {}).items():
            if s_num == 0:
                continue
            plex_s_data = plex_seasons_data.get(s_num)
            if plex_s_data and tvmaze_s_data and \
               plex_s_data.get('episodes_in_plex', 0) >= tvmaze_s_data.get('total_episodes', 0) and \
               tvmaze_s_data.get('total_episodes', 0) > 0:
                complete_plex_regular_seasons_count += 1
        series_status_text = f"{complete_plex_regular_seasons_count}/{tvmaze_total_regular_seasons}"

        if tvmaze_total_regular_seasons == 0:
            series_fill = STYLES['fills']['gray']
        elif complete_plex_regular_seasons_count >= tvmaze_total_regular_seasons:
            series_fill = STYLES['fills']['green']
        elif complete_plex_regular_seasons_count > 0:
            series_fill = STYLES['fills']['red']
        else:
            series_fill = None
    else:
        plex_regular_season_count = sum(1 for s_num in plex_seasons_data if s_num > 0)
        series_status_text = f"{plex_regular_season_count}/?"
        series_fill = STYLES['fills']['yellow'] if plex_regular_season_count > 0 else None

    return series_status_text, series_fill


def _calculate_season_cell(season_num: int, tvmaze_info: Optional[TVMazeShowInfo], plex_seasons_data: Dict[int, PlexSeasonData]) -> Tuple[str, Optional[PatternFill]]:
    """
    Calculates season cell text and fill color for a specific season.

    Args:
        season_num: Season number to calculate.
        tvmaze_info: TVMaze show information.
        plex_seasons_data: Plex season data.

    Returns:
        Tuple[str, Optional[PatternFill]]: A tuple of (season_text, fill_pattern)
    """
    season_text = ""
    season_fill = None
    plex_s_ep_count = plex_seasons_data.get(season_num, {}).get('episodes_in_plex', 0)

    if tvmaze_info and tvmaze_info.get('seasons'):
        tvmaze_s_data = tvmaze_info['seasons'].get(season_num)
        if tvmaze_s_data:
            tvmaze_s_ep_count = tvmaze_s_data.get('total_episodes', 0)
            season_text = f"{plex_s_ep_count}/{tvmaze_s_ep_count}"
            if tvmaze_s_ep_count == 0:
                season_fill = STYLES['fills']['gray'] if plex_s_ep_count == 0 else STYLES['fills']['yellow']
            elif plex_s_ep_count >= tvmaze_s_ep_count:
                season_fill = STYLES['fills']['green']
            elif plex_s_ep_count > 0:
                season_fill = STYLES['fills']['red']
        else:
            if plex_s_ep_count > 0:
                season_text, season_fill = f"{plex_s_ep_count}/?", STYLES['fills']['yellow']
            else:
                season_fill = STYLES['fills']['gray']
    else:
        if plex_s_ep_count > 0:
            season_text, season_fill = f"{plex_s_ep_count}/?", STYLES['fills']['yellow']
        elif season_num > max(plex_seasons_data.keys(), default=-1):
            season_fill = STYLES['fills']['gray']

    return season_text, season_fill


def _write_tv_show_row(ws: Any, row_idx: int, show_info_full: Dict[str, Any], season_headers_list: List[str]) -> None:
    """
    Writes a single TV show row with all its data.

    Args:
        ws: Worksheet object.
        row_idx: Row index to write to.
        show_info_full: Full show information dictionary.
        season_headers_list: List of season header names.
    """
    current_col = 1

    # Write base metadata fields
    for field_name in SELECTED_SHOW_FIELDS:
        value = show_info_full.get(field_name, 'N/A')
        cell = ws.cell(row=row_idx, column=current_col)
        cell.value = value if isinstance(value, (int, float)) and not pd.isna(value) else str(value if not pd.isna(value) else '')
        alignment_key = 'summary' if field_name in ['Summary', 'Tagline'] else \
                        'left' if field_name in ['Title', 'Studio', 'Collections', 'Genres', 'Labels'] else 'center'
        apply_cell_styling(cell, alignment_key=alignment_key)
        current_col += 1

    # Write series completion status
    tvmaze_info = show_info_full.get('tvmaze_info')
    plex_seasons_data = show_info_full.get('seasons_in_plex', {})
    series_status_text, series_fill = _calculate_series_completion(tvmaze_info, plex_seasons_data)
    status_cell = ws.cell(row=row_idx, column=current_col, value=series_status_text)
    apply_cell_styling(status_cell, fill_pattern=series_fill, alignment_key='center')
    current_col += 1

    # Write individual season cells
    for season_header_text in season_headers_list:
        season_num = int(season_header_text[1:])
        season_text, season_fill = _calculate_season_cell(season_num, tvmaze_info, plex_seasons_data)
        cell = ws.cell(row=row_idx, column=current_col)
        cell.value = season_text
        apply_cell_styling(cell, fill_pattern=season_fill, alignment_key='center')
        current_col += 1


def create_tv_shows_worksheet(section_name: str, wb: Workbook, shows_data_full: List[Dict[str, Any]], max_seasons_overall: int) -> None:
    """
    Creates and populates a worksheet for TV shows.

    Args:
        section_name: The name of the Plex library section.
        wb: The openpyxl Workbook object.
        shows_data_full: A list of dictionaries, each containing full data for a TV show.
        max_seasons_overall: The maximum number of seasons found across all shows for header generation.
    """
    if not shows_data_full:
        logger.info(f"No TV shows found in section '{section_name}'. Skipping sheet creation.")
        return

    safe_sheet_name = "".join(c if c.isalnum() or c in " _-" else "_" for c in section_name[:31])
    ws = wb.create_sheet(safe_sheet_name)

    # Sort the show data
    sort_by_field_show = 'Title' if 'Title' in SELECTED_SHOW_FIELDS else (SELECTED_SHOW_FIELDS[0] if SELECTED_SHOW_FIELDS else None)
    if sort_by_field_show:
        shows_data_full.sort(key=lambda x: str(x.get(sort_by_field_show, '')).lower())

    # Generate headers
    final_headers, season_headers_list = _generate_tv_show_headers(shows_data_full, max_seasons_overall)

    # Freeze panes
    ws.freeze_panes = get_column_letter(len(SELECTED_SHOW_FIELDS) + 1) + '2' if len(SELECTED_SHOW_FIELDS) > 0 else 'A2'

    # Write header row
    for col_idx, header_text in enumerate(final_headers, 1):
        apply_cell_styling(ws.cell(row=1, column=col_idx, value=header_text), is_header=True, alignment_key='center')

    # Write data rows
    for row_idx, show_info_full in enumerate(shows_data_full, 2):
        _write_tv_show_row(ws, row_idx, show_info_full, season_headers_list)

    # Adjust column widths
    auto_adjust_columns(ws, min_width=7, max_width=15, wrap_text_columns=['Summary', 'Tagline', 'Collections', 'Genres', 'Labels'])

    # Set specific widths for Title and Studio
    title_col_idx = SELECTED_SHOW_FIELDS.index('Title') + 1 if 'Title' in SELECTED_SHOW_FIELDS else None
    studio_col_idx = SELECTED_SHOW_FIELDS.index('Studio') + 1 if 'Studio' in SELECTED_SHOW_FIELDS else None
    if title_col_idx:
        ws.column_dimensions[get_column_letter(title_col_idx)].width = 35
    if studio_col_idx:
        ws.column_dimensions[get_column_letter(studio_col_idx)].width = 20

    # Create Excel table
    if len(shows_data_full) > 0 and len(final_headers) > 0:
        last_col_letter = get_column_letter(len(final_headers))
        table_name_base = "".join(c if c.isalnum() else "_" for c in safe_sheet_name)
        create_table(ws, f"{table_name_base}_Table", f"A1:{last_col_letter}{len(shows_data_full) + 1}") 

def check_file_writable(filename: str) -> bool:
    """
    Checks if a file is writable or if its directory is writable for creation.

    Args:
        filename: The path to the file to check.

    Returns:
        bool: True if the file can be written, False otherwise.
    """
    # Using walrus operator for Python 3.8+ to assign and check in one expression
    return (os.access(filename, os.W_OK) if os.path.exists(filename)
            else os.access((p_dir := os.path.dirname(filename) or '.'), os.W_OK))

def main() -> None:
    """Main function to orchestrate the Plex media export process."""
    logger.info("Plex Media Export Script started.")

    # Critical check: Ensure Plex URL and Token are provided.
    if not PLEX_URL or not PLEX_TOKEN:
        logger.error(f"Critical Error: {PLEX_URL_ENV_VAR} and/or {PLEX_TOKEN_ENV_VAR} are not set in your .env file or environment.")
        sys.exit(1)
        
    # Connect to Plex server.
    plex = connect_to_plex(); 
    if not plex: sys.exit(1) # Exit if connection failed.
    
    # Setup output filename with a timestamp.
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    base_filename = f"PlexMediaExport_{timestamp}.xlsx"
    output_dir = PLEX_EXPORT_DIR # Get output directory from config (defaults to current)
    
    # Ensure output directory exists, create if not.
    if not os.path.isdir(output_dir):
        try:
            os.makedirs(output_dir, exist_ok=True)
            logger.info(f"Created output directory: {os.path.abspath(output_dir)}")
        except OSError as e:
            logger.error(f"Could not create output directory {output_dir}: {e}")
            sys.exit(1)
            
    filename = os.path.join(output_dir, base_filename) # Full path to the output file.
    
    # Check if the output file is writable.
    if not check_file_writable(filename):
        logger.error(f"Cannot write to target file {filename}. Check permissions or path.")
        sys.exit(1)
        
    # Create a new Excel workbook and remove the default sheet.
    wb = Workbook(); 
    if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])
    
    total_movies_processed, total_shows_processed = 0, 0
    
    # Fetch library sections from Plex.
    try:
        logger.info("Fetching library sections from Plex...")
        sections = [s for s in plex.library.sections() if s.type in ['movie', 'show']] # Filter for movie or show sections
        logger.info(f"Found {len(sections)} movie/show sections.")
    except Exception as e:
        logger.error(f"Error fetching library sections from Plex: {e}")
        sys.exit(1)
        
    # If no relevant sections found, save an empty report and exit.
    if not sections:
        logger.warning("No movie or show sections found in your Plex library.")
        try:
            wb.save(filename)
            logger.info(f"Empty report saved as: {os.path.abspath(filename)}")
        except Exception as e:
            logger.error(f"Error saving empty workbook: {e}")
        sys.exit(0)
        
    # Process each section.
    for section in sections:
        logger.info(f"\nProcessing section: {section.title} (Type: {section.type})")
        try:
            items = section.all() # Get all items (movies or shows) in the current section.
            if not items:
                logger.info(f"  No items found in section '{section.title}'.")
                continue
            logger.info(f"  Found {len(items)} items in '{section.title}'.")
            
            # Process based on section type.
            if section.type == 'movie':
                logger.info(f"  Fetching details for {len(items)} movies...")
                movie_list = get_movie_details(items)
                if movie_list:
                    create_movies_worksheet(section.title, wb, movie_list)
                    total_movies_processed += len(movie_list)
                else:
                    logger.warning(f"  No movie details to write for section '{section.title}'.")
            elif section.type == 'show':
                logger.info(f"  Fetching details for {len(items)} TV shows...")
                shows_data, max_seasons = get_show_details(items)
                if shows_data:
                    create_tv_shows_worksheet(section.title, wb, shows_data, max_seasons)
                    total_shows_processed += len(shows_data)
                else:
                    logger.warning(f"  No TV show details to write for section '{section.title}'.")
        except Exception as e:
            logger.error(f"  An error occurred while processing section '{section.title}': {e}")
            # import traceback # Uncomment for full traceback during debugging
            # traceback.print_exc()

    # Save the final Excel workbook.
    try:
        wb.save(filename)
        logger.info(f"\n-----------------------------------------------------------")
        logger.info(f"Export complete!")
        logger.info(f"Processed {total_movies_processed} movies and {total_shows_processed} TV series.")
        logger.info(f"Report saved as: {os.path.abspath(filename)}")
        logger.info(f"-----------------------------------------------------------")
    except Exception as e:
        logger.error(f"Error saving workbook to {filename}: {e}")

    # Save the persistent TVMaze cache for future runs
    save_tvmaze_cache()

# Standard Python idiom: ensure main() is called only when the script is executed directly.
if __name__ == "__main__": 
    main()
# End of script
